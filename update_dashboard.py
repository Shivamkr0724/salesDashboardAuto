#!/usr/bin/env python3
"""
CAPS Onboarding Dashboard Updater
==================================
Fully automated: downloads Excel files from Google Drive,
builds the dashboard, pushes to GitHub → Vercel auto-deploys.

Run manually:   python update_dashboard.py
Scheduled:      via launchd (Mac) or Task Scheduler (Windows)

Setup:  pip install openpyxl requests gdown
"""

import os, sys, json, base64, webbrowser, shutil, time
from datetime import datetime
from collections import defaultdict

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION  ← edit these paths to match your machine
# ─────────────────────────────────────────────────────────────────────────────

# ═══════════════════════════════════════════════════════════════════
#  CONFIGURATION  ← Fill these in once, never touch again
# ═══════════════════════════════════════════════════════════════════

# ── GOOGLE DRIVE FILE IDs ─────────────────────────────────────────
# How to get a file ID:
#   1. Open the file in Google Drive
#   2. Click Share → Copy link
#   3. The ID is the long string between /d/ and /edit in the URL
#      e.g. https://drive.google.com/file/d/1ABC...XYZ/view
#                                           ^^^^^^^^^^^^ this part
#

GDRIVE_FILES = {
    "Onboarding_Master_.xlsx":                       "19rk0cduPsY-Yjbw8O9epgpcNB-biLmOEHEBu6PAKT3Y",
    "POA_Sheet.xlsx":                                "1YSfu3_g44wcrvDZw5TWp8thoA8GmroR0wnJ5QxpFHQU",
    "CRM_Sales_Caps__Responses_.xlsx":               "1f8ZvsIqsFzJeBzjJYATrMLreJJ5lYIqL-PEwYgV1WxI",
    "Data_Entry-_Onboarding_sheet__GMAP_Link_.xlsx": "1w7ut6Wbqn98VyoLZ5nS1bqQTf3r-BI6oEFZvlivqfKQ",
    "On_Boarding_Documents.xlsx":                    "1t66qxIC2d7CyvW9-Kcd0tp-l6zwcO24ewIP5vLD_4ng"
}
# If IDs are blank, script falls back to DATA_FOLDER below

# ── LOCAL FALLBACK (if not using Google Drive) ───────────────────
DATA_FOLDER = r"."   # folder containing the Excel files

# File names
FILE_MASTER  = "Onboarding_Master_.xlsx"
FILE_POA     = "POA_Sheet.xlsx"
FILE_CRM     = "CRM_Sales_Caps__Responses_.xlsx"
FILE_GMAP    = "Data_Entry-_Onboarding_sheet__GMAP_Link_.xlsx"
FILE_DOCS    = "On_Boarding_Documents.xlsx"

# Output
OUTPUT_HTML  = "caps_dashboard.html"

# ── GITHUB → VERCEL AUTO-DEPLOY ──────────────────────────────────
# Full path to your cloned GitHub repo on this Mac
# e.g. "/Users/macbook/Documents/caps-dashboard"
GITHUB_REPO_PATH  = "/Users/macbook/Desktop/salesDashboardAuto"
GITHUB_COMMIT_MSG = "dashboard: auto-update {date}"

# ── NOTIFICATION (optional) ──────────────────────────────────────
# Send a WhatsApp/email style Mac notification when done
SHOW_NOTIFICATION = True
LIVE_URL          = ""  # your Vercel URL e.g. "https://caps-dashboard.vercel.app"

# ── NETLIFY (leave blank if using Vercel) ────────────────────────
NETLIFY_SITE_ID = ""
NETLIFY_TOKEN   = ""

# ─────────────────────────────────────────────────────────────────────────────

def log(msg, ok=True):
    icon = "✅" if ok else "❌"
    print(f"{icon} {msg}")

def warn(msg):
    print(f"⚠️  {msg}")

def check_dependencies():
    missing = ["openpyxl", "requests"]
    # gdown needed for Google Drive downloads
    if any(v for v in GDRIVE_FILES.values()):
        missing.append("gdown")
    still_missing = []
    for pkg in missing:
        try:
            __import__(pkg)
        except ImportError:
            still_missing.append(pkg)
    if still_missing:
        print(f"\n📦 Installing packages: {', '.join(still_missing)}...\n")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install",
                               "--quiet"] + still_missing)
        print()

def download_from_gdrive():
    """Download all Excel files from Google Drive into a temp folder."""
    ids = {k: v for k, v in GDRIVE_FILES.items() if v.strip()}
    if not ids:
        return None  # no IDs configured, use local files

    try:
        import gdown
    except ImportError:
        warn("gdown not installed — skipping Google Drive download")
        return None

    tmp = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_gdrive_tmp")
    os.makedirs(tmp, exist_ok=True)

    print("\n☁️  Downloading from Google Drive...")
    all_ok = True
    for fname, file_id in ids.items():
        dest = os.path.join(tmp, fname)
        url  = f"https://drive.google.com/uc?id={file_id}"
        try:
            # Remove cached version first so we always get fresh file
            if os.path.exists(dest):
                os.remove(dest)
            gdown.download(url, dest, quiet=True)
            if os.path.exists(dest):
                size = os.path.getsize(dest) // 1024
                log(f"Downloaded {fname} ({size} KB)")
            else:
                warn(f"Failed to download {fname} — will use local copy if available")
                all_ok = False
        except Exception as e:
            warn(f"Error downloading {fname}: {e}")
            all_ok = False

    return tmp

def resolve(filename, gdrive_tmp=None):
    """Find file - checks gdrive tmp first, then DATA_FOLDER, then script dir."""
    paths = []
    if gdrive_tmp:
        paths.append(os.path.join(gdrive_tmp, filename))
    paths += [
        os.path.join(DATA_FOLDER, filename),
        filename,
        os.path.join(os.path.dirname(os.path.abspath(__file__)), filename),
    ]
    for p in paths:
        if os.path.exists(p):
            return p
    return None

def kam_name(email):
    if not email or str(email).strip() in ('', 'None', 'TEST', 'nan'): return None
    e = str(email).strip().lower()
    if '@' not in e:
        return e.strip().title() if len(e) > 2 else None
    return e.split('@')[0].replace('.', ' ').title()

def norm_date(d):
    if not d: return None
    try:
        if hasattr(d, 'date'): return str(d.date())
        s = str(d)[:10]
        if len(s) == 10 and s[4] == '-': return s
    except: pass
    return None

def load_excel(path, sheet=0):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True, keep_links=False)
        except Exception as e:
            print(f"  ⚠️  Warning reading {os.path.basename(path)}: {e}")
            return []
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        names = wb.sheetnames
        ws = wb[sheet] if sheet in names else wb.worksheets[0]
    # Stream rows directly — faster than list() for large files
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)
        # Stop after first fully-empty row to skip trailing blanks
    wb.close()
    return rows

# ─── EXTRACT DATA ────────────────────────────────────────────────────────────

def extract_master(path):
    rows = load_excel(path, 0)
    out = []
    for r in rows[1:]:
        if not r[0]: continue
        out.append({
            'dt':     norm_date(r[1]),
            'rest':   str(r[3] or '').strip(),
            'status': str(r[5] or '').strip(),
            'kam':    kam_name(r[6]),
            'loc':    str(r[13] or '').strip(),
            'zone':   str(r[14] or '').strip(),
            'cat':    str(r[8] or '').strip(),
        })
    return out

def extract_poa(path):
    rows = load_excel(path, 0)
    out = []
    for r in rows[1:]:
        if not r[0]: continue
        vd = norm_date(r[0])
        if not vd or vd < '2026-01-01': continue
        out.append({
            'vd':    vd,
            'kam':   kam_name(r[1]),
            'rm':    kam_name(r[2]),
            'rest':  str(r[3] or '').strip(),
            'lead':  str(r[4] or '').strip(),
            'loc':   str(r[6] or '').strip(),
            'zone':  str(r[7] or '').strip(),
            'vtype': str(r[8] or '').strip(),
            'de':    str(r[9] or '').strip(),
            'rem':   str(r[10] or '').strip(),
        })
    return out

def extract_crm(path):
    rows = load_excel(path, 0)
    out = []
    for r in rows[3:]:
        if not r[0]: continue
        vd = norm_date(r[1])
        if not vd or vd < '2026-01-01': continue
        rest = str(r[69] or r[67] or r[66] or r[4] or '').strip()
        loc  = str(r[65] or '').strip()
        if not loc and r[64]: loc = str(r[64])[:30].strip()
        out.append({
            'vd':  vd,
            'fd':  norm_date(r[49]),
            'mt':  str(r[3] or '').strip(),
            'rest':rest,
            'ob':  str(r[44] or '').strip(),
            'fu':  str(r[46] or '').strip(),
            'ag':  str(r[47] or '').strip(),
            'pts': str(r[48] or '').strip(),
            'kam': kam_name(r[51]),
            'poc': str(r[20] or r[54] or '').strip(),
            'ph':  str(r[21] or r[55] or '').strip(),
            'loc': loc,
            'gmap':str(r[18] or '').strip(),
            'kind':str(r[9] or '').strip(),
            'src': 'CRM',
        })
    return out

def extract_docs(path):
    rows = load_excel(path, 0)  # Sheet1
    out = []
    for r in rows[1:]:
        if not r[0]: continue
        out.append({
            'dt':     norm_date(r[0]),
            'rest':   str(r[1] or '').strip(),
            'kam':    kam_name(r[2]),
            'loc':    str(r[7] or '').strip(),
            'zone':   str(r[9] or '').strip(),
            'fssai':  str(r[24] or '').strip().lower(),
            'gst':    str(r[25] or '').strip().lower(),
            'pan':    str(r[26] or '').strip().lower(),
            'cheque': str(r[27] or '').strip().lower(),
            'menu':   str(r[28] or '').strip().lower(),
            'labour': str(r[29] or '').strip().lower(),
        })
    return out

def extract_whatsapp(path):
    # whatsapp sheet is index 1 in On_Boarding_Documents.xlsx
    try:
        rows = load_excel(path, 'whatsapp')
    except:
        rows = load_excel(path, 1)
    out = []
    for r in rows[1:]:
        if not r[0]: continue
        vd = norm_date(r[0])
        if not vd: continue
        out.append({
            'vd':   vd,
            'fd':   norm_date(r[7]),
            'rest': str(r[1] or '').strip(),
            'loc':  str(r[2] or '').strip(),
            'poc':  str(r[3] or '').strip(),
            'ph':   str(r[4] or '').strip(),
            'kam':  kam_name(r[5]),
            'com':  str(r[6] or '').strip(),
            'gmap': str(r[9] or '').strip(),
            'src':  'WA',
        })
    return out

def build_followups(crm, whatsapp):
    fu_crm = [r for r in crm if r['fu'].lower() == 'yes' and r['fd']]
    fu_wa  = [r for r in whatsapp if r['fd']]

    all_fu = []
    for r in fu_crm:
        all_fu.append({
            'fd':   r['fd'], 'vd': r['vd'], 'rest': r['rest'],
            'kam':  r['kam'], 'ag': r['ag'], 'pts': r['pts'],
            'poc':  r['poc'], 'ph': r['ph'],
            'loc':  r['loc'], 'gmap': r['gmap'],
            'com':  '', 'kind': r['kind'], 'src': 'CRM',
        })

    # Enrich with WhatsApp comments
    wa_map = defaultdict(list)
    for r in whatsapp:
        wa_map[r['rest'].lower()[:12]].append(r)

    for fu in all_fu:
        key = fu['rest'].lower()[:12]
        matches = sorted(wa_map.get(key, []), key=lambda x: x['vd'], reverse=True)
        if matches:
            fu['com'] = matches[0]['com']
            if not fu['loc']: fu['loc'] = matches[0]['loc']

    # Add WhatsApp-only follow-ups
    crm_keys = set(r['rest'].lower()[:12] for r in fu_crm)
    for r in fu_wa:
        if r['rest'].lower()[:12] not in crm_keys:
            all_fu.append({
                'fd': r['fd'], 'vd': r['vd'], 'rest': r['rest'],
                'kam': r['kam'], 'ag': '', 'pts': '',
                'poc': r['poc'], 'ph': r['ph'],
                'loc': r['loc'], 'gmap': r['gmap'],
                'com': r['com'], 'kind': '', 'src': 'WA',
            })

    return all_fu

# ─── MAC NOTIFICATION ────────────────────────────────────────────────────────

def notify_mac(title, message):
    """Show a Mac system notification."""
    if not SHOW_NOTIFICATION or sys.platform != 'darwin':
        return
    try:
        import subprocess as sp
        script = f'''display notification "{message}" with title "{title}" sound name "Glass"''' 
        sp.run(['osascript', '-e', script], capture_output=True)
    except Exception:
        pass  # Notifications are optional

# ─── GITHUB AUTO-PUSH (triggers Vercel deploy) ───────────────────────────────

def push_to_github(html_path, repo_path, commit_msg):
    """Copy dashboard to GitHub repo and push — Vercel picks it up automatically."""
    import subprocess as sp
    import shutil

    if not repo_path or not os.path.exists(repo_path):
        warn(f"GitHub repo path not found: {repo_path}")
        warn("  Set GITHUB_REPO_PATH in this script to enable auto-deploy.")
        return None

    print("\n🚀 Pushing to GitHub → Vercel will auto-deploy...")

    # Copy dashboard HTML as index.html into the repo
    dest = os.path.join(repo_path, 'index.html')
    shutil.copy2(html_path, dest)
    log(f"Copied dashboard to: {dest}")

    # Git commands
    def git(cmd, cwd=repo_path):
        result = sp.run(
            ['git'] + cmd, cwd=cwd,
            capture_output=True, text=True
        )
        return result.returncode, result.stdout.strip(), result.stderr.strip()

    # Check git is installed
    rc, out, err = git(['--version'], cwd=os.getcwd())
    if rc != 0:
        warn("Git not installed. Download from: https://git-scm.com/downloads")
        return None

    # Stage the file
    rc, out, err = git(['add', 'index.html'])
    if rc != 0:
        warn(f"git add failed: {err}")
        return None

    # Check if there's anything to commit
    rc, out, err = git(['status', '--porcelain'])
    if not out.strip():
        log("No changes to commit — dashboard is already up to date")
        # Get the live URL from git remote
        rc2, remote, _ = git(['remote', 'get-url', 'origin'])
        return remote

    # Commit
    msg = commit_msg.replace('{date}', datetime.now().strftime('%d %b %Y %H:%M'))
    rc, out, err = git(['commit', '-m', msg])
    if rc != 0:
        warn(f"git commit failed: {err}")
        return None
    log(f"Committed: {msg}")

    # Push
    rc, out, err = git(['push'])
    if rc != 0:
        warn(f"git push failed: {err}")
        warn("  Make sure you've run 'git clone' and have push access to the repo.")
        return None

    log("Pushed to GitHub ✅")
    log("Vercel will deploy in ~10 seconds automatically")

    # Return the Vercel URL if we can guess it
    rc, remote_url, _ = git(['remote', 'get-url', 'origin'])
    return remote_url

# ─── NETLIFY DEPLOY ──────────────────────────────────────────────────────────

def deploy_netlify(html_path):
    import requests, zipfile, io
    print("\n🚀 Deploying to Netlify...")

    # Zip the HTML as index.html
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(html_path, 'index.html')
    buf.seek(0)

    url = f"https://api.netlify.com/api/v1/sites/{NETLIFY_SITE_ID}/deploys"
    headers = {
        "Authorization": f"Bearer {NETLIFY_TOKEN}",
        "Content-Type": "application/zip",
    }
    resp = requests.post(url, headers=headers, data=buf.read(), timeout=60)

    if resp.status_code in (200, 201):
        data = resp.json()
        site_url = data.get('ssl_url') or data.get('url', '')
        log(f"Deployed! URL: {site_url}")
        return site_url
    else:
        warn(f"Netlify deploy failed: {resp.status_code} {resp.text[:200]}")
        return None

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  CAPS Onboarding Dashboard Updater")
    print(f"  {datetime.now().strftime('%d %b %Y  %H:%M')}")
    print("=" * 55)

    check_dependencies()

    # ── Download from Google Drive (if IDs configured) ──────────────────
    gdrive_tmp = download_from_gdrive()

    # ── Check files ────────────────────────────────────────────────────────
    files = {
        'Master':  FILE_MASTER,
        'POA':     FILE_POA,
        'CRM':     FILE_CRM,
        'GMap':    FILE_GMAP,
        'Docs':    FILE_DOCS,
    }
    paths = {}
    all_found = True
    for label, fname in files.items():
        p = resolve(fname, gdrive_tmp)
        if p:
            src = "☁️  GDrive" if gdrive_tmp and gdrive_tmp in p else "💾 Local"
            log(f"Found {label} [{src}]: {os.path.basename(p)}")
            paths[label] = p
        else:
            warn(f"{label} file not found: {fname}")
            all_found = False

    if not all_found:
        print("\n❌ Some files missing.")
        if not any(v for v in GDRIVE_FILES.values()):
            print("   Set GDRIVE_FILES IDs in this script for auto-download.")
        print("   Or place Excel files in:", os.path.abspath(DATA_FOLDER))
        notify_mac("❌ CAPS Dashboard", "Update failed — Excel files not found")
        input("\nPress Enter to close...")
        return

    print("\n📊 Processing data...")
    import time
    t0 = time.time()

    print("  Reading Onboarding Master...", end=" ", flush=True)
    master = extract_master(paths['Master'])
    print(f"{len(master)} rows ✅")

    print("  Reading POA Sheet...", end=" ", flush=True)
    poa = extract_poa(paths['POA'])
    print(f"{len(poa)} rows ✅")

    print("  Reading CRM (largest file)...", end=" ", flush=True)
    crm = extract_crm(paths['CRM'])
    print(f"{len(crm)} rows ✅")

    print("  Reading Onboarding Documents...", end=" ", flush=True)
    docs = extract_docs(paths['Docs'])
    print(f"{len(docs)} rows ✅")

    print("  Reading Visit Log (WhatsApp)...", end=" ", flush=True)
    whatsapp = extract_whatsapp(paths['Docs'])
    print(f"{len(whatsapp)} rows ✅")

    print("  Building follow-up list...", end=" ", flush=True)
    fu = build_followups(crm, whatsapp)
    print(f"{len(fu)} entries ✅")

    print(f"\n  ⏱  Data processed in {time.time()-t0:.1f} seconds")

    # Serialize
    data_json = json.dumps({
        'crm': crm, 'poa': poa, 'wa': whatsapp,
        'master': master, 'docs': docs, 'fu': fu,
    }, default=str, ensure_ascii=False)

    print("\n🔨 Building dashboard HTML...")

    # Decode template
    tmpl_before = base64.b64decode("PCFET0NUWVBFIGh0bWw+CjxodG1sIGxhbmc9ImVuIj4KPGhlYWQ+CjxtZXRhIGNoYXJzZXQ9IlVURi04Ij4KPG1ldGEgbmFtZT0idmlld3BvcnQiIGNvbnRlbnQ9IndpZHRoPWRldmljZS13aWR0aCwgaW5pdGlhbC1zY2FsZT0xLjAiPgo8dGl0bGU+Q0FQUyDigJMgT25ib2FyZGluZyBJbnRlbGxpZ2VuY2U8L3RpdGxlPgo8bGluayBocmVmPSJodHRwczovL2ZvbnRzLmdvb2dsZWFwaXMuY29tL2NzczI/ZmFtaWx5PURNK1NhbnM6aXRhbCx3Z2h0QDAsMzAwOzAsNDAwOzAsNTAwOzAsNjAwOzEsNDAwJmZhbWlseT1CZWJhcytOZXVlJmRpc3BsYXk9c3dhcCIgcmVsPSJzdHlsZXNoZWV0Ij4KPHNjcmlwdCBzcmM9Imh0dHBzOi8vY2RuanMuY2xvdWRmbGFyZS5jb20vYWpheC9saWJzL0NoYXJ0LmpzLzQuNC4xL2NoYXJ0LnVtZC5qcyI+PC9zY3JpcHQ+CjxzdHlsZT4KOnJvb3R7CiAgLS1iZzojMGMwZjE4Oy0tc3VyZjojMTMxODI2Oy0tY2FyZDojMWEyMDMwOy0tY2FyZDI6IzFlMjUzODsKICAtLWIxOiMyNDJkNDI7LS1iMjojMmQzYTUyOy0tYjM6IzNhNGE2MDsKICAtLXR4OiNkZGUzZjI7LS1tdTojN2E4N2E4Oy0tZGltOiMzYTQ0NjA7CiAgLS1hbTojZjVhNjIzOy0tdGw6IzE0YjhhNjstLWJsOiM0YTkwZDk7CiAgLS1nbjojMjJjNTVlOy0tcnM6I2Y0M2Y1ZTstLXZpOiM4YjVjZjY7LS1vcjojZjk3MzE2Oy0tY3k6IzA2YjZkNDsKICAtLXI6MTJweDsKfQoqe2JveC1zaXppbmc6Ym9yZGVyLWJveDttYXJnaW46MDtwYWRkaW5nOjB9Cmh0bWx7c2Nyb2xsLWJlaGF2aW9yOnNtb290aH0KYm9keXtmb250LWZhbWlseTonRE0gU2Fucycsc2Fucy1zZXJpZjtiYWNrZ3JvdW5kOnZhcigtLWJnKTtjb2xvcjp2YXIoLS10eCk7Zm9udC1zaXplOjEzcHg7bWluLWhlaWdodDoxMDB2aDtkaXNwbGF5OmZsZXh9CgovKiDilIDilIAgU0lERUJBUiDilIDilIAgKi8KLnNie3dpZHRoOjIwMHB4O21pbi1oZWlnaHQ6MTAwdmg7YmFja2dyb3VuZDp2YXIoLS1zdXJmKTtib3JkZXItcmlnaHQ6MXB4IHNvbGlkIHZhcigtLWIxKTtkaXNwbGF5OmZsZXg7ZmxleC1kaXJlY3Rpb246Y29sdW1uO2ZsZXgtc2hyaW5rOjA7cG9zaXRpb246c3RpY2t5O3RvcDowO2hlaWdodDoxMDB2aDtvdmVyZmxvdy15OmF1dG99Ci5zYi1sb2dve3BhZGRpbmc6MTZweCAxNHB4IDE0cHg7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgdmFyKC0tYjEpfQouc2ItbG9nby1yb3d7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6OXB4fQouc2ItaWNvbnt3aWR0aDozMHB4O2hlaWdodDozMHB4O2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyx2YXIoLS1hbSksdmFyKC0tcnMpKTtib3JkZXItcmFkaXVzOjdweDtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2p1c3RpZnktY29udGVudDpjZW50ZXI7Zm9udC1zaXplOjE0cHg7ZmxleC1zaHJpbms6MH0KLnNiLWJyYW5kIGgxe2ZvbnQtZmFtaWx5OidCZWJhcyBOZXVlJztmb250LXNpemU6MS4xcmVtO2xldHRlci1zcGFjaW5nOjFweDtjb2xvcjp2YXIoLS10eCl9Ci5zYi1icmFuZCBwe2ZvbnQtc2l6ZTouNThyZW07Y29sb3I6dmFyKC0tbXUpO21hcmdpbi10b3A6MXB4fQouc2Itc2Vje3BhZGRpbmc6MTJweCAwIDJweH0KLnNiLWxibHtmb250LXNpemU6LjU4cmVtO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtsZXR0ZXItc3BhY2luZzouMTJlbTtjb2xvcjp2YXIoLS1kaW0pO3BhZGRpbmc6MCAxNHB4O21hcmdpbi1ib3R0b206M3B4fQouc2ItaXRlbXtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo4cHg7cGFkZGluZzo3cHggMTRweDtjb2xvcjp2YXIoLS1tdSk7Zm9udC1zaXplOi43OHJlbTtmb250LXdlaWdodDo1MDA7Y3Vyc29yOnBvaW50ZXI7Ym9yZGVyLWxlZnQ6MnB4IHNvbGlkIHRyYW5zcGFyZW50O3RyYW5zaXRpb246YWxsIC4xMnM7dXNlci1zZWxlY3Q6bm9uZX0KLnNiLWl0ZW06aG92ZXJ7YmFja2dyb3VuZDp2YXIoLS1jYXJkKTtjb2xvcjp2YXIoLS10eCl9Ci5zYi1pdGVtLm9ue2JhY2tncm91bmQ6dmFyKC0tY2FyZCk7Y29sb3I6dmFyKC0tYW0pO2JvcmRlci1sZWZ0LWNvbG9yOnZhcigtLWFtKX0KLnNiLWljb3t3aWR0aDoxNXB4O3RleHQtYWxpZ246Y2VudGVyO2ZvbnQtc2l6ZToxM3B4fQouc2ItZm9vdHttYXJnaW4tdG9wOmF1dG87cGFkZGluZzoxMnB4IDE0cHg7Ym9yZGVyLXRvcDoxcHggc29saWQgdmFyKC0tYjEpO2ZvbnQtc2l6ZTouNjVyZW07Y29sb3I6dmFyKC0tbXUpfQouc2ItZm9vdCBzdHJvbmd7ZGlzcGxheTpibG9jaztjb2xvcjp2YXIoLS10eCk7bWFyZ2luLWJvdHRvbToycHg7Zm9udC1zaXplOi42OHJlbX0KLnVwZHtkaXNwbGF5OmlubGluZS1ibG9jazttYXJnaW4tdG9wOjVweDtwYWRkaW5nOjJweCA4cHg7YmFja2dyb3VuZDpyZ2JhKDM0LDE5Nyw5NCwuMSk7Y29sb3I6dmFyKC0tZ24pO2JvcmRlcjoxcHggc29saWQgcmdiYSgzNCwxOTcsOTQsLjIpO2JvcmRlci1yYWRpdXM6MjBweDtmb250LXNpemU6LjZyZW19CgovKiDilIDilIAgTUFJTiDilIDilIAgKi8KLm1haW57ZmxleDoxO21pbi13aWR0aDowO2Rpc3BsYXk6ZmxleDtmbGV4LWRpcmVjdGlvbjpjb2x1bW59CgovKiDilIDilIAgVE9QQkFSIOKUgOKUgCAqLwoudG9wYmFye2JhY2tncm91bmQ6dmFyKC0tc3VyZik7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgdmFyKC0tYjEpO3BhZGRpbmc6MTBweCAyMnB4O2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjEwcHg7cG9zaXRpb246c3RpY2t5O3RvcDowO3otaW5kZXg6NTA7ZmxleC13cmFwOndyYXA7Z2FwOjhweH0KLnRvcGJhci10aXRsZXtmb250LWZhbWlseTonQmViYXMgTmV1ZSc7Zm9udC1zaXplOjEuMXJlbTtsZXR0ZXItc3BhY2luZzouNXB4O3doaXRlLXNwYWNlOm5vd3JhcH0KLnRvcGJhci1zdWJ7Zm9udC1zaXplOi42OHJlbTtjb2xvcjp2YXIoLS1tdSk7bWFyZ2luLWxlZnQ6NHB4fQoudGItcmlnaHR7bWFyZ2luLWxlZnQ6YXV0bztkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo2cHg7ZmxleC13cmFwOndyYXB9CgovKiBEQVRFIFJBTkdFIEZJTFRFUiAqLwouZHItd3JhcHtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo1cHg7YmFja2dyb3VuZDp2YXIoLS1jYXJkKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWIyKTtib3JkZXItcmFkaXVzOjhweDtwYWRkaW5nOjRweCA4cHh9Ci5kci13cmFwIGxhYmVse2ZvbnQtc2l6ZTouNjVyZW07Y29sb3I6dmFyKC0tbXUpO3doaXRlLXNwYWNlOm5vd3JhcH0KLmRyLXdyYXAgaW5wdXRbdHlwZT1kYXRlXXtiYWNrZ3JvdW5kOnRyYW5zcGFyZW50O2JvcmRlcjpub25lO2NvbG9yOnZhcigtLXR4KTtmb250LWZhbWlseTonRE0gU2Fucyc7Zm9udC1zaXplOi43MnJlbTtvdXRsaW5lOm5vbmU7Y3Vyc29yOnBvaW50ZXJ9Ci5kci1zZXB7Y29sb3I6dmFyKC0tbXUpO2ZvbnQtc2l6ZTouN3JlbX0KLmRyLXByZXNldHN7ZGlzcGxheTpmbGV4O2dhcDo0cHh9Ci5wci1idG57YmFja2dyb3VuZDp2YXIoLS1jYXJkMik7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iMSk7Y29sb3I6dmFyKC0tbXUpO2JvcmRlci1yYWRpdXM6NnB4O3BhZGRpbmc6NHB4IDlweDtmb250LWZhbWlseTonRE0gU2Fucyc7Zm9udC1zaXplOi43cmVtO2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246YWxsIC4xMnM7d2hpdGUtc3BhY2U6bm93cmFwfQoucHItYnRuOmhvdmVyLC5wci1idG4ub257YmFja2dyb3VuZDp2YXIoLS1hbSk7Y29sb3I6IzAwMDtib3JkZXItY29sb3I6dmFyKC0tYW0pO2ZvbnQtd2VpZ2h0OjYwMH0KCi5zZWx7YmFja2dyb3VuZDp2YXIoLS1jYXJkKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWIyKTtjb2xvcjp2YXIoLS10eCk7Zm9udC1mYW1pbHk6J0RNIFNhbnMnO2ZvbnQtc2l6ZTouNzJyZW07cGFkZGluZzo1cHggOXB4O2JvcmRlci1yYWRpdXM6N3B4O291dGxpbmU6bm9uZTtjdXJzb3I6cG9pbnRlcn0KLnNlbCBvcHRpb257YmFja2dyb3VuZDp2YXIoLS1jYXJkKX0KLnRhYi1idG57YmFja2dyb3VuZDp2YXIoLS1jYXJkMik7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iMSk7Y29sb3I6dmFyKC0tbXUpO2JvcmRlci1yYWRpdXM6N3B4O3BhZGRpbmc6NXB4IDEycHg7Zm9udC1mYW1pbHk6J0RNIFNhbnMnO2ZvbnQtc2l6ZTouNzNyZW07Y3Vyc29yOnBvaW50ZXI7dHJhbnNpdGlvbjphbGwgLjEyczt3aGl0ZS1zcGFjZTpub3dyYXB9Ci50YWItYnRuOmhvdmVye2NvbG9yOnZhcigtLXR4KTtib3JkZXItY29sb3I6dmFyKC0tYjIpfQoudGFiLWJ0bi5vbntiYWNrZ3JvdW5kOnZhcigtLWFtKTtjb2xvcjojMDAwO2JvcmRlci1jb2xvcjp2YXIoLS1hbSk7Zm9udC13ZWlnaHQ6NjAwfQoKLyog4pSA4pSAIENPTlRFTlQg4pSA4pSAICovCi5jb250ZW50e3BhZGRpbmc6MThweCAyMnB4IDQwcHg7ZmxleDoxfQoucGFuZWx7ZGlzcGxheTpub25lfQoucGFuZWwub257ZGlzcGxheTpibG9ja30KCi8qIOKUgOKUgCBGSUxURVIgQkFSIOKUgOKUgCAqLwouZmJhcntkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo4cHg7bWFyZ2luLWJvdHRvbToxNnB4O2ZsZXgtd3JhcDp3cmFwO2JhY2tncm91bmQ6dmFyKC0tY2FyZCk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iMSk7Ym9yZGVyLXJhZGl1czp2YXIoLS1yKTtwYWRkaW5nOjEwcHggMTRweH0KLmZiYXItbGFiZWx7Zm9udC1zaXplOi42NXJlbTtjb2xvcjp2YXIoLS1tdSk7d2hpdGUtc3BhY2U6bm93cmFwfQoKLyog4pSA4pSAIEtQSSBST1cg4pSA4pSAICovCi5rcm93e2Rpc3BsYXk6Z3JpZDtncmlkLXRlbXBsYXRlLWNvbHVtbnM6cmVwZWF0KDUsMWZyKTtnYXA6MTBweDttYXJnaW4tYm90dG9tOjE0cHh9Ci5rcm93NHtncmlkLXRlbXBsYXRlLWNvbHVtbnM6cmVwZWF0KDQsMWZyKX0KLmtjYXJke2JhY2tncm91bmQ6dmFyKC0tY2FyZCk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iMSk7Ym9yZGVyLXJhZGl1czp2YXIoLS1yKTtwYWRkaW5nOjE0cHggMTZweDtwb3NpdGlvbjpyZWxhdGl2ZTtvdmVyZmxvdzpoaWRkZW47Y3Vyc29yOmRlZmF1bHR9Ci5rY2FyZDo6YWZ0ZXJ7Y29udGVudDonJztwb3NpdGlvbjphYnNvbHV0ZTtib3R0b206MDtsZWZ0OjA7cmlnaHQ6MDtoZWlnaHQ6MnB4fQoua2MxOjphZnRlcntiYWNrZ3JvdW5kOnZhcigtLWFtKX0gLmtjMjo6YWZ0ZXJ7YmFja2dyb3VuZDp2YXIoLS1nbil9Ci5rYzM6OmFmdGVye2JhY2tncm91bmQ6dmFyKC0tYmwpfSAua2M0OjphZnRlcntiYWNrZ3JvdW5kOnZhcigtLXJzKX0KLmtjNTo6YWZ0ZXJ7YmFja2dyb3VuZDp2YXIoLS12aSl9Ci5rbGJse2ZvbnQtc2l6ZTouNjJyZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2xldHRlci1zcGFjaW5nOi4wOGVtO2NvbG9yOnZhcigtLW11KTttYXJnaW4tYm90dG9tOjZweH0KLmt2YWx7Zm9udC1mYW1pbHk6J0JlYmFzIE5ldWUnO2ZvbnQtc2l6ZToxLjlyZW07bGV0dGVyLXNwYWNpbmc6LjNweDtjb2xvcjp2YXIoLS10eCk7bGluZS1oZWlnaHQ6MX0KLmtzdWJ7Zm9udC1zaXplOi42NHJlbTtjb2xvcjp2YXIoLS1tdSk7bWFyZ2luLXRvcDo0cHh9Ci5rcGlsbHtkaXNwbGF5OmlubGluZS1ibG9jaztmb250LXNpemU6LjZyZW07Zm9udC13ZWlnaHQ6NjAwO3BhZGRpbmc6MXB4IDdweDtib3JkZXItcmFkaXVzOjIwcHg7bWFyZ2luLXRvcDo0cHh9Ci5wZ3tiYWNrZ3JvdW5kOnJnYmEoMzQsMTk3LDk0LC4xKTtjb2xvcjp2YXIoLS1nbil9Ci5wcntiYWNrZ3JvdW5kOnJnYmEoMjQ0LDYzLDk0LC4xKTtjb2xvcjp2YXIoLS1ycyl9Ci5wYXtiYWNrZ3JvdW5kOnJnYmEoMjQ1LDE2NiwzNSwuMSk7Y29sb3I6dmFyKC0tYW0pfQoucGJ7YmFja2dyb3VuZDpyZ2JhKDc0LDE0NCwyMTcsLjEpO2NvbG9yOnZhcigtLWJsKX0KLnB2e2JhY2tncm91bmQ6cmdiYSgxMzksOTIsMjQ2LC4xKTtjb2xvcjp2YXIoLS12aSl9CgovKiDilIDilIAgU0VDVElPTiDilIDilIAgKi8KLnNlYy10aXRsZXtmb250LWZhbWlseTonQmViYXMgTmV1ZSc7Zm9udC1zaXplOi45cmVtO2xldHRlci1zcGFjaW5nOi41cHg7Y29sb3I6dmFyKC0tbXUpO21hcmdpbi1ib3R0b206MTBweDtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo4cHh9Ci5zZWMtdGl0bGU6OmFmdGVye2NvbnRlbnQ6Jyc7ZmxleDoxO2hlaWdodDoxcHg7YmFja2dyb3VuZDp2YXIoLS1iMSl9CgovKiDilIDilIAgR1JJRCDilIDilIAgKi8KLmcye2Rpc3BsYXk6Z3JpZDtncmlkLXRlbXBsYXRlLWNvbHVtbnM6MWZyIDFmcjtnYXA6MTJweDttYXJnaW4tYm90dG9tOjEycHh9Ci5nM3tkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOjFmciAxZnIgMWZyO2dhcDoxMnB4O21hcmdpbi1ib3R0b206MTJweH0KLmZ1bGx7Z3JpZC1jb2x1bW46MS8tMX0KCi8qIOKUgOKUgCBDSEFSVCBDQVJEIOKUgOKUgCAqLwouY2N7YmFja2dyb3VuZDp2YXIoLS1jYXJkKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWIxKTtib3JkZXItcmFkaXVzOnZhcigtLXIpO3BhZGRpbmc6MTZweCAxOHB4fQouY2MtaGVhZHttYXJnaW4tYm90dG9tOjEycHh9Ci5jYy10aXRsZXtmb250LXNpemU6LjgzcmVtO2ZvbnQtd2VpZ2h0OjYwMDtjb2xvcjp2YXIoLS10eCl9Ci5jYy1zdWJ7Zm9udC1zaXplOi42NXJlbTtjb2xvcjp2YXIoLS1tdSk7bWFyZ2luLXRvcDoycHh9Ci5jd3twb3NpdGlvbjpyZWxhdGl2ZTt3aWR0aDoxMDAlfQoubGVne2Rpc3BsYXk6ZmxleDtmbGV4LXdyYXA6d3JhcDtnYXA6OHB4O21hcmdpbi10b3A6MTBweH0KLmxpe2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjRweDtmb250LXNpemU6LjY1cmVtO2NvbG9yOnZhcigtLW11KX0KLmxzcXt3aWR0aDo4cHg7aGVpZ2h0OjhweDtib3JkZXItcmFkaXVzOjJweDtmbGV4LXNocmluazowfQoKLyog4pSA4pSAIEhCQVIg4pSA4pSAICovCi5oYmFyLWl0ZW17ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6OHB4O21hcmdpbi1ib3R0b206N3B4fQouaGJhci1uYW1le2ZvbnQtc2l6ZTouNzJyZW07Y29sb3I6dmFyKC0tdHgpO2ZsZXgtc2hyaW5rOjA7d2hpdGUtc3BhY2U6bm93cmFwO292ZXJmbG93OmhpZGRlbjt0ZXh0LW92ZXJmbG93OmVsbGlwc2lzfQouaGJhci10cmFja3tmbGV4OjE7aGVpZ2h0OjE4cHg7YmFja2dyb3VuZDp2YXIoLS1zdXJmKTtib3JkZXItcmFkaXVzOjNweDtvdmVyZmxvdzpoaWRkZW59Ci5oYmFyLWZpbGx7aGVpZ2h0OjEwMCU7Ym9yZGVyLXJhZGl1czozcHg7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtwYWRkaW5nOjAgNnB4O2ZvbnQtc2l6ZTouNjJyZW07Zm9udC13ZWlnaHQ6NjAwO2NvbG9yOiNmZmY7d2hpdGUtc3BhY2U6bm93cmFwfQouaGJhci1udW17Zm9udC1zaXplOi43NXJlbTtmb250LXdlaWdodDo2MDA7Y29sb3I6dmFyKC0tdHgpO2ZsZXgtc2hyaW5rOjA7dGV4dC1hbGlnbjpyaWdodDttaW4td2lkdGg6MjhweH0KCi8qIOKUgOKUgCBGVU5ORUwg4pSA4pSAICovCi5mbi1yb3d7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6OHB4O21hcmdpbi1ib3R0b206OHB4fQouZm4tbGFiZWx7d2lkdGg6MTQwcHg7Zm9udC1zaXplOi43MnJlbTtjb2xvcjp2YXIoLS1tdSk7dGV4dC1hbGlnbjpyaWdodDtmbGV4LXNocmluazowO3doaXRlLXNwYWNlOm5vd3JhcDtvdmVyZmxvdzpoaWRkZW47dGV4dC1vdmVyZmxvdzplbGxpcHNpc30KLmZuLXRyYWNre2ZsZXg6MTtoZWlnaHQ6MzBweDtiYWNrZ3JvdW5kOnZhcigtLXN1cmYpO2JvcmRlci1yYWRpdXM6NHB4O292ZXJmbG93OmhpZGRlbn0KLmZuLWZpbGx7aGVpZ2h0OjEwMCU7Ym9yZGVyLXJhZGl1czo0cHg7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtqdXN0aWZ5LWNvbnRlbnQ6ZmxleC1lbmQ7cGFkZGluZzowIDEwcHg7Zm9udC1zaXplOi43cmVtO2ZvbnQtd2VpZ2h0OjcwMDtjb2xvcjojZmZmO3RyYW5zaXRpb246d2lkdGggLjVzfQouZm4tY291bnR7d2lkdGg6NDVweDt0ZXh0LWFsaWduOnJpZ2h0O2ZvbnQtc2l6ZTouNzVyZW07Zm9udC13ZWlnaHQ6NzAwO2NvbG9yOnZhcigtLXR4KTtmbGV4LXNocmluazowfQoKLyog4pSA4pSAIERPQyBCQVJTIOKUgOKUgCAqLwouZG9jLXJvd3tiYWNrZ3JvdW5kOnZhcigtLWNhcmQyKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWIxKTtib3JkZXItcmFkaXVzOjhweDtwYWRkaW5nOjEwcHggMTJweDttYXJnaW4tYm90dG9tOjhweH0KLmRvYy1oZHtkaXNwbGF5OmZsZXg7anVzdGlmeS1jb250ZW50OnNwYWNlLWJldHdlZW47bWFyZ2luLWJvdHRvbTo1cHg7Zm9udC1zaXplOi43cmVtfQouZG9jLW5te2ZvbnQtd2VpZ2h0OjYwMDtjb2xvcjp2YXIoLS10eCl9Ci5kb2MtcGN0e2NvbG9yOnZhcigtLWFtKTtmb250LXdlaWdodDo3MDB9Ci5kb2MtYmFye2hlaWdodDo1cHg7YmFja2dyb3VuZDp2YXIoLS1iMSk7Ym9yZGVyLXJhZGl1czozcHg7b3ZlcmZsb3c6aGlkZGVufQouZG9jLWZpbGx7aGVpZ2h0OjEwMCU7Ym9yZGVyLXJhZGl1czozcHh9CgovKiDilIDilIAgREFZIFBMQU4gQ0FSRFMg4pSA4pSAICovCi5kYXlwbGFuLWdyaWR7ZGlzcGxheTpncmlkO2dyaWQtdGVtcGxhdGUtY29sdW1uczpyZXBlYXQoYXV0by1maWxsLG1pbm1heCgzMDBweCwxZnIpKTtnYXA6MTBweH0KLmRwLWNhcmR7YmFja2dyb3VuZDp2YXIoLS1jYXJkKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWIxKTtib3JkZXItcmFkaXVzOnZhcigtLXIpO3BhZGRpbmc6MTRweCAxNnB4O3RyYW5zaXRpb246Ym9yZGVyLWNvbG9yIC4xNXN9Ci5kcC1jYXJkOmhvdmVye2JvcmRlci1jb2xvcjp2YXIoLS1iMyl9Ci5kcC10b3B7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmZsZXgtc3RhcnQ7anVzdGlmeS1jb250ZW50OnNwYWNlLWJldHdlZW47bWFyZ2luLWJvdHRvbTo4cHh9Ci5kcC1yZXN0e2ZvbnQtc2l6ZTouODJyZW07Zm9udC13ZWlnaHQ6NjAwO2NvbG9yOnZhcigtLXR4KTtsaW5lLWhlaWdodDoxLjN9Ci5kcC1rYW17Zm9udC1zaXplOi42M3JlbTtjb2xvcjp2YXIoLS1tdSk7bWFyZ2luLXRvcDoycHh9Ci5kcC1iYWRnZXN7ZGlzcGxheTpmbGV4O2dhcDo0cHg7ZmxleC13cmFwOndyYXB9Ci5iYWRnZXtkaXNwbGF5OmlubGluZS1ibG9jaztmb250LXNpemU6LjZyZW07Zm9udC13ZWlnaHQ6NzAwO3BhZGRpbmc6MnB4IDdweDtib3JkZXItcmFkaXVzOjIwcHh9Ci5iLWhvdHtiYWNrZ3JvdW5kOnJnYmEoMjQ0LDYzLDk0LC4xMik7Y29sb3I6dmFyKC0tcnMpfQouYi13YXJte2JhY2tncm91bmQ6cmdiYSgyNDUsMTY2LDM1LC4xMik7Y29sb3I6dmFyKC0tYW0pfQouYi1jb2xke2JhY2tncm91bmQ6cmdiYSg3NCwxNDQsMjE3LC4xMik7Y29sb3I6dmFyKC0tYmwpfQouYi1jcm17YmFja2dyb3VuZDpyZ2JhKDEzOSw5MiwyNDYsLjEyKTtjb2xvcjp2YXIoLS12aSl9Ci5iLXdhe2JhY2tncm91bmQ6cmdiYSgyMCwxODQsMTY2LC4xMik7Y29sb3I6dmFyKC0tdGwpfQouYi1vbmJ7YmFja2dyb3VuZDpyZ2JhKDM0LDE5Nyw5NCwuMTIpO2NvbG9yOnZhcigtLWduKX0KLmRwLXJvd3tkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6ZmxleC1zdGFydDtnYXA6NnB4O21hcmdpbi10b3A6NXB4O2ZvbnQtc2l6ZTouN3JlbTtjb2xvcjp2YXIoLS1tdSk7bGluZS1oZWlnaHQ6MS40fQouZHAtcm93IC5pY297Zm9udC1zaXplOjEycHg7ZmxleC1zaHJpbms6MDttYXJnaW4tdG9wOjFweH0KLmRwLXJvdyAudmFse2NvbG9yOnZhcigtLXR4KX0KLmRwLXJvdyBhe2NvbG9yOnZhcigtLWJsKTt0ZXh0LWRlY29yYXRpb246bm9uZX0KLmRwLXJvdyBhOmhvdmVye3RleHQtZGVjb3JhdGlvbjp1bmRlcmxpbmV9Ci5kcC1jb21tZW50e2JhY2tncm91bmQ6dmFyKC0tc3VyZik7Ym9yZGVyLWxlZnQ6MnB4IHNvbGlkIHZhcigtLWFtKTtib3JkZXItcmFkaXVzOjAgNHB4IDRweCAwO3BhZGRpbmc6NHB4IDhweDtmb250LXNpemU6LjY4cmVtO2NvbG9yOnZhcigtLW11KTttYXJnaW4tdG9wOjZweDtmb250LXN0eWxlOml0YWxpY30KLmRwLWVtcHR5e3RleHQtYWxpZ246Y2VudGVyO3BhZGRpbmc6NDBweDtjb2xvcjp2YXIoLS1tdSk7Zm9udC1zaXplOi44MnJlbX0KCi8qIOKUgOKUgCBLQU0gRlVOTkVMIFRBQkxFIOKUgOKUgCAqLwoua2YtdGFibGV7d2lkdGg6MTAwJTtib3JkZXItY29sbGFwc2U6Y29sbGFwc2U7Zm9udC1zaXplOi43NXJlbX0KLmtmLXRhYmxlIHRoe3RleHQtYWxpZ246bGVmdDtwYWRkaW5nOjhweCAxMHB4O2ZvbnQtc2l6ZTouNjJyZW07dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2xldHRlci1zcGFjaW5nOi4wOGVtO2NvbG9yOnZhcigtLW11KTtib3JkZXItYm90dG9tOjFweCBzb2xpZCB2YXIoLS1iMSk7Zm9udC13ZWlnaHQ6NTAwO3doaXRlLXNwYWNlOm5vd3JhcH0KLmtmLXRhYmxlIHRke3BhZGRpbmc6OHB4IDEwcHg7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgzNiw0NSw2NiwuNSk7dmVydGljYWwtYWxpZ246bWlkZGxlfQoua2YtdGFibGUgdHI6bGFzdC1jaGlsZCB0ZHtib3JkZXItYm90dG9tOm5vbmV9Ci5rZi10YWJsZSB0cjpob3ZlciB0ZHtiYWNrZ3JvdW5kOnJnYmEoMjU1LDI1NSwyNTUsLjAyKX0KLnN0YWdlLWNlbGx7dGV4dC1hbGlnbjpjZW50ZXJ9Ci5zdGFnZS1waWxse2Rpc3BsYXk6aW5saW5lLWJsb2NrO3BhZGRpbmc6MnB4IDlweDtib3JkZXItcmFkaXVzOjIwcHg7Zm9udC1zaXplOi42NXJlbTtmb250LXdlaWdodDo3MDA7bWluLXdpZHRoOjI4cHg7dGV4dC1hbGlnbjpjZW50ZXJ9Ci5zLWludHJve2JhY2tncm91bmQ6cmdiYSg3NCwxNDQsMjE3LC4xMik7Y29sb3I6dmFyKC0tYmwpfQoucy1mdXtiYWNrZ3JvdW5kOnJnYmEoMjQ1LDE2NiwzNSwuMTIpO2NvbG9yOnZhcigtLWFtKX0KLnMtY3N7YmFja2dyb3VuZDpyZ2JhKDEzOSw5MiwyNDYsLjEyKTtjb2xvcjp2YXIoLS12aSl9Ci5zLWRje2JhY2tncm91bmQ6cmdiYSg2LDE4MiwyMTIsLjEyKTtjb2xvcjp2YXIoLS1jeSl9Ci5zLW9ie2JhY2tncm91bmQ6cmdiYSgzNCwxOTcsOTQsLjEyKTtjb2xvcjp2YXIoLS1nbil9Ci5yYW5rLWJhZGdle2Rpc3BsYXk6aW5saW5lLWZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2p1c3RpZnktY29udGVudDpjZW50ZXI7d2lkdGg6MTlweDtoZWlnaHQ6MTlweDtib3JkZXItcmFkaXVzOjUwJTtmb250LXNpemU6LjYycmVtO2ZvbnQtd2VpZ2h0OjcwMH0KLnJrMXtiYWNrZ3JvdW5kOnJnYmEoMjQ1LDE2NiwzNSwuMik7Y29sb3I6dmFyKC0tYW0pfQoucmsye2JhY2tncm91bmQ6cmdiYSgxNDgsMTYzLDE4NCwuMTUpO2NvbG9yOiM5NGEzYjh9Ci5yazN7YmFja2dyb3VuZDpyZ2JhKDE4MCwxMjAsNjAsLjE1KTtjb2xvcjojYjQ3YzNjfQoucmtue2JhY2tncm91bmQ6dmFyKC0tc3VyZik7Y29sb3I6dmFyKC0tbXUpfQoKLyog4pSA4pSAIFdBIFRBQkxFIOKUgOKUgCAqLwoud2EtdGFibGV7d2lkdGg6MTAwJTtib3JkZXItY29sbGFwc2U6Y29sbGFwc2U7Zm9udC1zaXplOi43NXJlbX0KLndhLXRhYmxlIHRoe3RleHQtYWxpZ246bGVmdDtwYWRkaW5nOjhweCAxMHB4O2ZvbnQtc2l6ZTouNnJlbTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7bGV0dGVyLXNwYWNpbmc6LjA4ZW07Y29sb3I6dmFyKC0tbXUpO2JvcmRlci1ib3R0b206MXB4IHNvbGlkIHZhcigtLWIxKTt3aGl0ZS1zcGFjZTpub3dyYXB9Ci53YS10YWJsZSB0ZHtwYWRkaW5nOjhweCAxMHB4O2JvcmRlci1ib3R0b206MXB4IHNvbGlkIHJnYmEoMzYsNDUsNjYsLjQpO2NvbG9yOnZhcigtLXR4KX0KLndhLXRhYmxlIHRyOmhvdmVyIHRke2JhY2tncm91bmQ6cmdiYSgyNTUsMjU1LDI1NSwuMDIpfQoud2EtdGFibGUgdHI6bGFzdC1jaGlsZCB0ZHtib3JkZXItYm90dG9tOm5vbmV9CgovKiDilIDilIAgUFJPR1JFU1Mg4pSA4pSAICovCi5wcm9nLXdyYXB7aGVpZ2h0OjZweDtiYWNrZ3JvdW5kOnZhcigtLWIxKTtib3JkZXItcmFkaXVzOjNweDtvdmVyZmxvdzpoaWRkZW59Ci5wcm9nLWZpbGx7aGVpZ2h0OjEwMCU7Ym9yZGVyLXJhZGl1czozcHh9CgovKiDilIDilIAgU0NST0xMQUJMRSDilIDilIAgKi8KLnNjcm9sbC14e292ZXJmbG93LXg6YXV0b30KLnNjcm9sbC15e292ZXJmbG93LXk6YXV0bzttYXgtaGVpZ2h0OjQyMHB4fQoKLyog4pSA4pSAIFJFU1BPTlNJVkUg4pSA4pSAICovCkBtZWRpYShtYXgtd2lkdGg6OTAwcHgpewogIC5zYntkaXNwbGF5Om5vbmV9CiAgLmtyb3d7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOnJlcGVhdCgyLDFmcil9CiAgLmcyLC5nM3tncmlkLXRlbXBsYXRlLWNvbHVtbnM6MWZyfQogIC5mdWxse2dyaWQtY29sdW1uOjF9CiAgLnRvcGJhcntmbGV4LXdyYXA6d3JhcH0KICAudGItcmlnaHR7d2lkdGg6MTAwJX0KfQo8L3N0eWxlPgo8L2hlYWQ+Cjxib2R5PgoKPCEtLSDilIDilIAgU0lERUJBUiDilIDilIAgLS0+CjxkaXYgY2xhc3M9InNiIj4KICA8ZGl2IGNsYXNzPSJzYi1sb2dvIj4KICAgIDxkaXYgY2xhc3M9InNiLWxvZ28tcm93Ij4KICAgICAgPGRpdiBjbGFzcz0ic2ItaWNvbiI+8J+NvTwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJzYi1icmFuZCI+PGgxPkNBUFM8L2gxPjxwPk9uYm9hcmRpbmcgSW50ZWxsaWdlbmNlPC9wPjwvZGl2PgogICAgPC9kaXY+CiAgPC9kaXY+CgogIDxkaXYgY2xhc3M9InNiLXNlYyI+CiAgICA8ZGl2IGNsYXNzPSJzYi1sYmwiPlBlcmZvcm1hbmNlPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzYi1pdGVtIG9uIiBvbmNsaWNrPSJnbygnb3ZlcnZpZXcnKSI+PHNwYW4gY2xhc3M9InNiLWljbyI+4qyhPC9zcGFuPk92ZXJ2aWV3PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzYi1pdGVtIiBvbmNsaWNrPSJnbygnbG9jYXRpb24nKSI+PHNwYW4gY2xhc3M9InNiLWljbyI+8J+TjTwvc3Bhbj5Mb2NhdGlvbiAmIFpvbmU8L2Rpdj4KICAgIDxkaXYgY2xhc3M9InNiLWl0ZW0iIG9uY2xpY2s9ImdvKCdrYW0nKSI+PHNwYW4gY2xhc3M9InNiLWljbyI+8J+RpDwvc3Bhbj5LQU0gV2lzZTwvZGl2PgogICAgPGRpdiBjbGFzcz0ic2ItaXRlbSIgb25jbGljaz0iZ28oJ2NvbnRyYWN0cycpIj48c3BhbiBjbGFzcz0ic2ItaWNvIj7wn5OLPC9zcGFuPkNvbnRyYWN0czwvZGl2PgogICAgPGRpdiBjbGFzcz0ic2ItaXRlbSIgb25jbGljaz0iZ28oJ2RvY3VtZW50cycpIj48c3BhbiBjbGFzcz0ic2ItaWNvIj7wn5OBPC9zcGFuPkRvY3VtZW50czwvZGl2PgogIDwvZGl2PgoKICA8ZGl2IGNsYXNzPSJzYi1zZWMiPgogICAgPGRpdiBjbGFzcz0ic2ItbGJsIj5PcGVyYXRpb25zPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzYi1pdGVtIiBvbmNsaWNrPSJnbygnZGF5cGxhbicpIj48c3BhbiBjbGFzcz0ic2ItaWNvIj7wn5eTPC9zcGFuPkRheSBQbGFuPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzYi1pdGVtIiBvbmNsaWNrPSJnbygncG9hJykiPjxzcGFuIGNsYXNzPSJzYi1pY28iPvCfk4U8L3NwYW4+UE9BIFRyYWNrZXI8L2Rpdj4KICAgIDxkaXYgY2xhc3M9InNiLWl0ZW0iIG9uY2xpY2s9ImdvKCdmdW5uZWwnKSI+PHNwYW4gY2xhc3M9InNiLWljbyI+4peRPC9zcGFuPktBTSBGdW5uZWw8L2Rpdj4KICAgIDxkaXYgY2xhc3M9InNiLWl0ZW0iIG9uY2xpY2s9ImdvKCd2aXNpdHMnKSI+PHNwYW4gY2xhc3M9InNiLWljbyI+8J+Xkjwvc3Bhbj5WaXNpdCBMb2c8L2Rpdj4KICA8L2Rpdj4KCiAgPGRpdiBjbGFzcz0ic2ItZm9vdCI+CiAgICA8c3Ryb25nPjUgRGF0YSBTb3VyY2VzPC9zdHJvbmc+CiAgICBPbmJvYXJkaW5nIE1hc3RlciDCtyBQT0EgwrcgQ1JNIMK3IERvY3MgwrcgVmlzaXQgTG9nCiAgICA8c3BhbiBjbGFzcz0idXBkIj5VcGRhdGVkIG9uIHVwbG9hZDwvc3Bhbj4KICA8L2Rpdj4KPC9kaXY+Cgo8IS0tIOKUgOKUgCBNQUlOIOKUgOKUgCAtLT4KPGRpdiBjbGFzcz0ibWFpbiI+Cgo8IS0tIOKUgOKUgCBUT1BCQVIg4pSA4pSAIC0tPgo8ZGl2IGNsYXNzPSJ0b3BiYXIiPgogIDxkaXY+CiAgICA8c3BhbiBjbGFzcz0idG9wYmFyLXRpdGxlIj5PbmJvYXJkaW5nIEludGVsbGlnZW5jZTwvc3Bhbj4KICAgIDxzcGFuIGNsYXNzPSJ0b3BiYXItc3ViIiBpZD0idGItc3ViIj5DQVBTIERlbGl2ZXJ5IMK3IEh5ZGVyYWJhZDwvc3Bhbj4KICA8L2Rpdj4KICA8ZGl2IGNsYXNzPSJ0Yi1yaWdodCI+CiAgICA8IS0tIERhdGUgcmFuZ2UgLS0+CiAgICA8ZGl2IGNsYXNzPSJkci13cmFwIj4KICAgICAgPGxhYmVsPkZyb208L2xhYmVsPgogICAgICA8aW5wdXQgdHlwZT0iZGF0ZSIgaWQ9ImRyLWZyb20iIHZhbHVlPSIyMDI2LTAyLTAxIiBvbmNoYW5nZT0iYXBwbHlGaWx0ZXJzKCkiPgogICAgICA8c3BhbiBjbGFzcz0iZHItc2VwIj7ihpI8L3NwYW4+CiAgICAgIDxpbnB1dCB0eXBlPSJkYXRlIiBpZD0iZHItdG8iIHZhbHVlPSIyMDI2LTAzLTE4IiBvbmNoYW5nZT0iYXBwbHlGaWx0ZXJzKCkiPgogICAgPC9kaXY+CiAgICA8IS0tIFByZXNldHMgLS0+CiAgICA8ZGl2IGNsYXNzPSJkci1wcmVzZXRzIj4KICAgICAgPGJ1dHRvbiBjbGFzcz0icHItYnRuIiBvbmNsaWNrPSJzZXRQcmVzZXQoJ3RvZGF5JykiPlRvZGF5PC9idXR0b24+CiAgICAgIDxidXR0b24gY2xhc3M9InByLWJ0biIgb25jbGljaz0ic2V0UHJlc2V0KCd5ZXN0ZXJkYXknKSI+WWVzdGVyZGF5PC9idXR0b24+CiAgICAgIDxidXR0b24gY2xhc3M9InByLWJ0biIgb25jbGljaz0ic2V0UHJlc2V0KCc3ZCcpIj43IERheXM8L2J1dHRvbj4KICAgICAgPGJ1dHRvbiBjbGFzcz0icHItYnRuIG9uIiBvbmNsaWNrPSJzZXRQcmVzZXQoJ210ZCcpIj5NVEQ8L2J1dHRvbj4KICAgICAgPGJ1dHRvbiBjbGFzcz0icHItYnRuIiBvbmNsaWNrPSJzZXRQcmVzZXQoJ2FsbCcpIj5BbGw8L2J1dHRvbj4KICAgIDwvZGl2PgogICAgPCEtLSBLQU0gZmlsdGVyIC0tPgogICAgPHNlbGVjdCBjbGFzcz0ic2VsIiBpZD0ic2VsLWthbSIgb25jaGFuZ2U9ImFwcGx5RmlsdGVycygpIj4KICAgICAgPG9wdGlvbiB2YWx1ZT0iYWxsIj5BbGwgS0FNczwvb3B0aW9uPgogICAgPC9zZWxlY3Q+CiAgICA8IS0tIFpvbmUgZmlsdGVyIC0tPgogICAgPHNlbGVjdCBjbGFzcz0ic2VsIiBpZD0ic2VsLXpvbmUiIG9uY2hhbmdlPSJhcHBseUZpbHRlcnMoKSI+CiAgICAgIDxvcHRpb24gdmFsdWU9ImFsbCI+QWxsIFpvbmVzPC9vcHRpb24+CiAgICA8L3NlbGVjdD4KICA8L2Rpdj4KPC9kaXY+Cgo8IS0tIOKUgOKUgCBDT05URU5UIOKUgOKUgCAtLT4KPGRpdiBjbGFzcz0iY29udGVudCI+Cgo8IS0tIOKVkOKVkCBPVkVSVklFVyDilZDilZAgLS0+CjxkaXYgY2xhc3M9InBhbmVsIG9uIiBpZD0icC1vdmVydmlldyI+CiAgPGRpdiBjbGFzcz0ia3JvdyIgaWQ9Im92LWtwaXMiPjwvZGl2PgogIDxkaXYgY2xhc3M9ImcyIj4KICAgIDxkaXYgY2xhc3M9ImNjIj4KICAgICAgPGRpdiBjbGFzcz0iY2MtaGVhZCI+PGRpdiBjbGFzcz0iY2MtdGl0bGUiPlpvbmUtd2lzZSBPbmJvYXJkaW5nczwvZGl2PjxkaXYgY2xhc3M9ImNjLXN1YiI+RnJvbSBPbmJvYXJkaW5nIERvY3MgwrcgZmlsdGVyZWQgcmFuZ2U8L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBpZD0ib3Ytem9uZS1iYXJzIj48L2Rpdj4KICAgIDwvZGl2PgogICAgPGRpdiBjbGFzcz0iY2MiPgogICAgICA8ZGl2IGNsYXNzPSJjYy1oZWFkIj48ZGl2IGNsYXNzPSJjYy10aXRsZSI+Q29udHJhY3QgU3RhdHVzPC9kaXY+PGRpdiBjbGFzcz0iY2Mtc3ViIj5Gcm9tIE9uYm9hcmRpbmcgTWFzdGVyPC9kaXY+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImN3IiBzdHlsZT0iaGVpZ2h0OjE4MHB4Ij48Y2FudmFzIGlkPSJvdi1jb250cmFjdCI+PC9jYW52YXM+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImxlZyIgaWQ9Im92LWNvbnRyYWN0LWxlZyI+PC9kaXY+CiAgICA8L2Rpdj4KICA8L2Rpdj4KICA8ZGl2IGNsYXNzPSJnMiI+CiAgICA8ZGl2IGNsYXNzPSJjYyI+CiAgICAgIDxkaXYgY2xhc3M9ImNjLWhlYWQiPjxkaXYgY2xhc3M9ImNjLXRpdGxlIj5DUk0gRnVubmVsIChPdmVydmlldyk8L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPkFsbCB2aXNpdCBzdGFnZXMgaW4gc2VsZWN0ZWQgcGVyaW9kPC9kaXY+PC9kaXY+CiAgICAgIDxkaXYgaWQ9Im92LWZ1bm5lbCI+PC9kaXY+CiAgICA8L2Rpdj4KICAgIDxkaXYgY2xhc3M9ImNjIj4KICAgICAgPGRpdiBjbGFzcz0iY2MtaGVhZCI+PGRpdiBjbGFzcz0iY2MtdGl0bGUiPkRvY3VtZW50IENvbGxlY3Rpb24gUmF0ZTwvZGl2PjxkaXYgY2xhc3M9ImNjLXN1YiI+Q29sbGVjdGVkIHZzIFBlbmRpbmcgwrcgYWxsIDYgZG9jczwvZGl2PjwvZGl2PgogICAgICA8ZGl2IGlkPSJvdi1kb2NzIj48L2Rpdj4KICAgIDwvZGl2PgogIDwvZGl2Pgo8L2Rpdj4KCjwhLS0g4pWQ4pWQIExPQ0FUSU9OIOKVkOKVkCAtLT4KPGRpdiBjbGFzcz0icGFuZWwiIGlkPSJwLWxvY2F0aW9uIj4KICA8ZGl2IGNsYXNzPSJnMiI+CiAgICA8ZGl2IGNsYXNzPSJjYyI+CiAgICAgIDxkaXYgY2xhc3M9ImNjLWhlYWQiPjxkaXYgY2xhc3M9ImNjLXRpdGxlIj5ab25lLXdpc2UgT25ib2FyZGluZ3M8L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPk9uYm9hcmRpbmcgRG9jdW1lbnRzIMK3IHpvbmUgZmllbGQ8L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY3ciIHN0eWxlPSJoZWlnaHQ6MjYwcHgiPjxjYW52YXMgaWQ9ImxvYy16b25lIj48L2NhbnZhcz48L2Rpdj4KICAgIDwvZGl2PgogICAgPGRpdiBjbGFzcz0iY2MiPgogICAgICA8ZGl2IGNsYXNzPSJjYy1oZWFkIj48ZGl2IGNsYXNzPSJjYy10aXRsZSI+UE9BIEFjdGl2aXR5IGJ5IFpvbmU8L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPkZvbGxvdy11cHMgKyBXYWxrLWlucyBmcm9tIFBPQSBTaGVldDwvZGl2PjwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJjdyIgc3R5bGU9ImhlaWdodDoyNjBweCI+PGNhbnZhcyBpZD0ibG9jLXBvYS16b25lIj48L2NhbnZhcz48L2Rpdj4KICAgIDwvZGl2PgogIDwvZGl2PgogIDxkaXYgY2xhc3M9ImNjIiBzdHlsZT0ibWFyZ2luLWJvdHRvbToxMnB4Ij4KICAgIDxkaXYgY2xhc3M9ImNjLWhlYWQiPjxkaXYgY2xhc3M9ImNjLXRpdGxlIj5Ub3AgTG9jYXRpb25zIGJ5IE9uYm9hcmRpbmdzPC9kaXY+PGRpdiBjbGFzcz0iY2Mtc3ViIj5Mb2NhbGl0eSBmaWVsZCBmcm9tIE9uYm9hcmRpbmcgRG9jdW1lbnRzPC9kaXY+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJjdyIgc3R5bGU9ImhlaWdodDozNDBweCI+PGNhbnZhcyBpZD0ibG9jLWxvY3MiPjwvY2FudmFzPjwvZGl2PgogIDwvZGl2Pgo8L2Rpdj4KCjwhLS0g4pWQ4pWQIEtBTSDilZDilZAgLS0+CjxkaXYgY2xhc3M9InBhbmVsIiBpZD0icC1rYW0iPgogIDxkaXYgY2xhc3M9ImcyIj4KICAgIDxkaXYgY2xhc3M9ImNjIj4KICAgICAgPGRpdiBjbGFzcz0iY2MtaGVhZCI+PGRpdiBjbGFzcz0iY2MtdGl0bGUiPk9uYm9hcmRpbmdzIGJ5IEtBTTwvZGl2PjxkaXYgY2xhc3M9ImNjLXN1YiI+T25ib2FyZGluZyBEb2NzIMK3IGZpbHRlcmVkIGRhdGUgcmFuZ2U8L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY3ciIHN0eWxlPSJoZWlnaHQ6MzAwcHgiPjxjYW52YXMgaWQ9ImthbS1iYXIiPjwvY2FudmFzPjwvZGl2PgogICAgPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJjYyI+CiAgICAgIDxkaXYgY2xhc3M9ImNjLWhlYWQiPjxkaXYgY2xhc3M9ImNjLXRpdGxlIj5LQU0gU2hhcmU8L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPiUgb2YgdG90YWwgb25ib2FyZGluZ3MgaW4gcGVyaW9kPC9kaXY+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImN3IiBzdHlsZT0iaGVpZ2h0OjI2MHB4Ij48Y2FudmFzIGlkPSJrYW0tZG9udXQiPjwvY2FudmFzPjwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJsZWciIGlkPSJrYW0tZGxlZyI+PC9kaXY+CiAgICA8L2Rpdj4KICA8L2Rpdj4KICA8ZGl2IGNsYXNzPSJjYyI+CiAgICA8ZGl2IGNsYXNzPSJjYy1oZWFkIj48ZGl2IGNsYXNzPSJjYy10aXRsZSI+S0FNIExlYWRlcmJvYXJkPC9kaXY+PGRpdiBjbGFzcz0iY2Mtc3ViIj5PbmJvYXJkaW5ncyDCtyBjb250cmFjdHMgwrcgZG9jIGNvbXBsZXRpb24gaW4gc2VsZWN0ZWQgcmFuZ2U8L2Rpdj48L2Rpdj4KICAgIDx0YWJsZSBjbGFzcz0ia2YtdGFibGUiIGlkPSJrYW0tbGIiPjwvdGFibGU+CiAgPC9kaXY+CjwvZGl2PgoKPCEtLSDilZDilZAgQ09OVFJBQ1RTIOKVkOKVkCAtLT4KPGRpdiBjbGFzcz0icGFuZWwiIGlkPSJwLWNvbnRyYWN0cyI+CiAgPGRpdiBjbGFzcz0ia3JvdyBrcm93NCIgaWQ9ImN0LWtwaXMiPjwvZGl2PgogIDxkaXYgY2xhc3M9ImcyIj4KICAgIDxkaXYgY2xhc3M9ImNjIj4KICAgICAgPGRpdiBjbGFzcz0iY2MtaGVhZCI+PGRpdiBjbGFzcz0iY2MtdGl0bGUiPkNvbnRyYWN0IFN0YXR1cyBCcmVha2Rvd248L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPk9uYm9hcmRpbmcgTWFzdGVyIMK3IHN0YXR1cyBjb2x1bW48L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY3ciIHN0eWxlPSJoZWlnaHQ6MjYwcHgiPjxjYW52YXMgaWQ9ImN0LWJhciI+PC9jYW52YXM+PC9kaXY+CiAgICA8L2Rpdj4KICAgIDxkaXYgY2xhc3M9ImNjIj4KICAgICAgPGRpdiBjbGFzcz0iY2MtaGVhZCI+PGRpdiBjbGFzcz0iY2MtdGl0bGUiPktBTS13aXNlIENvbnRyYWN0IFJhdGU8L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPlN1Ym1pdHRlZCB2cyBQZW5kaW5nIHBlciBLQU08L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY3ciIHN0eWxlPSJoZWlnaHQ6MjYwcHgiPjxjYW52YXMgaWQ9ImN0LWthbSI+PC9jYW52YXM+PC9kaXY+CiAgICA8L2Rpdj4KICA8L2Rpdj4KPC9kaXY+Cgo8IS0tIOKVkOKVkCBET0NVTUVOVFMg4pWQ4pWQIC0tPgo8ZGl2IGNsYXNzPSJwYW5lbCIgaWQ9InAtZG9jdW1lbnRzIj4KICA8ZGl2IGNsYXNzPSJrcm93IiBpZD0iZG9jLWtwaXMiPjwvZGl2PgogIDxkaXYgY2xhc3M9ImcyIj4KICAgIDxkaXYgY2xhc3M9ImNjIj4KICAgICAgPGRpdiBjbGFzcz0iY2MtaGVhZCI+PGRpdiBjbGFzcz0iY2MtdGl0bGUiPkNvbGxlY3RlZCB2cyBQZW5kaW5nPC9kaXY+PGRpdiBjbGFzcz0iY2Mtc3ViIj5PbmJvYXJkaW5nIERvY3MgwrcgYWxsIDYgZG9jdW1lbnQgdHlwZXM8L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY3ciIHN0eWxlPSJoZWlnaHQ6MjYwcHgiPjxjYW52YXMgaWQ9ImRvYy1iYXIiPjwvY2FudmFzPjwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJsZWciPjxzcGFuIGNsYXNzPSJsaSI+PHNwYW4gY2xhc3M9ImxzcSIgc3R5bGU9ImJhY2tncm91bmQ6dmFyKC0tZ24pIj48L3NwYW4+Q29sbGVjdGVkPC9zcGFuPjxzcGFuIGNsYXNzPSJsaSI+PHNwYW4gY2xhc3M9ImxzcSIgc3R5bGU9ImJhY2tncm91bmQ6dmFyKC0tcnMpIj48L3NwYW4+UGVuZGluZzwvc3Bhbj48L2Rpdj4KICAgIDwvZGl2PgogICAgPGRpdiBjbGFzcz0iY2MiPgogICAgICA8ZGl2IGNsYXNzPSJjYy1oZWFkIj48ZGl2IGNsYXNzPSJjYy10aXRsZSI+RG9jdW1lbnQgQ29tcGxldGlvbiBEZXRhaWw8L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPiUgY29sbGVjdGVkIHBlciBkb2N1bWVudCB0eXBlPC9kaXY+PC9kaXY+CiAgICAgIDxkaXYgaWQ9ImRvYy1kZXRhaWwiPjwvZGl2PgogICAgPC9kaXY+CiAgPC9kaXY+CjwvZGl2PgoKPCEtLSDilZDilZAgREFZIFBMQU4g4pWQ4pWQIC0tPgo8ZGl2IGNsYXNzPSJwYW5lbCIgaWQ9InAtZGF5cGxhbiI+CiAgPGRpdiBjbGFzcz0iY2MiIHN0eWxlPSJtYXJnaW4tYm90dG9tOjE0cHgiPgogICAgPGRpdiBzdHlsZT0iZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6MTBweDtmbGV4LXdyYXA6d3JhcCI+CiAgICAgIDxkaXY+CiAgICAgICAgPGRpdiBjbGFzcz0iY2MtdGl0bGUiPvCfk4UgTWFuYWdlciBEYXkgUGxhbjwvZGl2PgogICAgICAgIDxkaXYgY2xhc3M9ImNjLXN1YiI+UmVzdGF1cmFudHMgd2l0aCBmb2xsb3ctdXAgZHVlIGluIHNlbGVjdGVkIGRhdGUgcmFuZ2UgwrcgZnJvbSBDUk0gKyBXaGF0c0FwcCBsb2c8L2Rpdj4KICAgICAgPC9kaXY+CiAgICAgIDxkaXYgc3R5bGU9Im1hcmdpbi1sZWZ0OmF1dG87ZGlzcGxheTpmbGV4O2dhcDo2cHg7ZmxleC13cmFwOndyYXAiPgogICAgICAgIDxzZWxlY3QgY2xhc3M9InNlbCIgaWQ9ImRwLWthbSIgb25jaGFuZ2U9InJlbmRlckRheVBsYW4oKSI+PG9wdGlvbiB2YWx1ZT0iYWxsIj5BbGwgS0FNczwvb3B0aW9uPjwvc2VsZWN0PgogICAgICAgIDxzZWxlY3QgY2xhc3M9InNlbCIgaWQ9ImRwLXNyYyIgb25jaGFuZ2U9InJlbmRlckRheVBsYW4oKSI+CiAgICAgICAgICA8b3B0aW9uIHZhbHVlPSJhbGwiPkFsbCBTb3VyY2VzPC9vcHRpb24+CiAgICAgICAgICA8b3B0aW9uIHZhbHVlPSJDUk0iPkNSTSBPbmx5PC9vcHRpb24+CiAgICAgICAgICA8b3B0aW9uIHZhbHVlPSJXQSI+V2hhdHNBcHAgT25seTwvb3B0aW9uPgogICAgICAgIDwvc2VsZWN0PgogICAgICAgIDxzZWxlY3QgY2xhc3M9InNlbCIgaWQ9ImRwLWFnIiBvbmNoYW5nZT0icmVuZGVyRGF5UGxhbigpIj4KICAgICAgICAgIDxvcHRpb24gdmFsdWU9ImFsbCI+QWxsIEFnZW5kYXM8L29wdGlvbj4KICAgICAgICAgIDxvcHRpb24gdmFsdWU9IkNvbnRyYWN0IFNpZ25pbmciPkNvbnRyYWN0IFNpZ25pbmc8L29wdGlvbj4KICAgICAgICAgIDxvcHRpb24gdmFsdWU9IkRvY3VtZW50IENvbGxlY3Rpb24iPkRvY3VtZW50IENvbGxlY3Rpb248L29wdGlvbj4KICAgICAgICAgIDxvcHRpb24gdmFsdWU9Ik1lZXRpbmcgd2l0aCBkaWZmZXJlbnQgUE9DIj5NZWV0aW5nIHcvIERpZmZlcmVudCBQT0M8L29wdGlvbj4KICAgICAgICA8L3NlbGVjdD4KICAgICAgPC9kaXY+CiAgICA8L2Rpdj4KICAgIDxkaXYgaWQ9ImRwLXN1bW1hcnkiIHN0eWxlPSJkaXNwbGF5OmZsZXg7Z2FwOjEwcHg7bWFyZ2luLXRvcDoxMnB4O2ZsZXgtd3JhcDp3cmFwIj48L2Rpdj4KICA8L2Rpdj4KICA8ZGl2IGNsYXNzPSJkYXlwbGFuLWdyaWQiIGlkPSJkcC1jYXJkcyI+PC9kaXY+CjwvZGl2PgoKPCEtLSDilZDilZAgUE9BIFRSQUNLRVIg4pWQ4pWQIC0tPgo8ZGl2IGNsYXNzPSJwYW5lbCIgaWQ9InAtcG9hIj4KICA8ZGl2IGNsYXNzPSJrcm93IGtyb3c0IiBpZD0icG9hLWtwaXMiPjwvZGl2PgogIDxkaXYgY2xhc3M9ImcyIj4KICAgIDxkaXYgY2xhc3M9ImNjIj4KICAgICAgPGRpdiBjbGFzcz0iY2MtaGVhZCI+PGRpdiBjbGFzcz0iY2MtdGl0bGUiPkRhaWx5IEZvbGxvdy11cHMgdnMgV2Fsay1pbnM8L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPlBPQSBTaGVldCDCtyB2aXNpdCB0eXBlIHBlciBkYXk8L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY3ciIHN0eWxlPSJoZWlnaHQ6MjIwcHgiPjxjYW52YXMgaWQ9InBvYS1kYWlseSI+PC9jYW52YXM+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImxlZyI+CiAgICAgICAgPHNwYW4gY2xhc3M9ImxpIj48c3BhbiBjbGFzcz0ibHNxIiBzdHlsZT0iYmFja2dyb3VuZDp2YXIoLS1hbSkiPjwvc3Bhbj5Gb2xsb3ctdXA8L3NwYW4+CiAgICAgICAgPHNwYW4gY2xhc3M9ImxpIj48c3BhbiBjbGFzcz0ibHNxIiBzdHlsZT0iYmFja2dyb3VuZDp2YXIoLS1ibCkiPjwvc3Bhbj5XYWxrLWluPC9zcGFuPgogICAgICAgIDxzcGFuIGNsYXNzPSJsaSI+PHNwYW4gY2xhc3M9ImxzcSIgc3R5bGU9ImJhY2tncm91bmQ6dmFyKC0tZ24pIj48L3NwYW4+T25ib2FyZGluZzwvc3Bhbj4KICAgICAgPC9kaXY+CiAgICA8L2Rpdj4KICAgIDxkaXYgY2xhc3M9ImNjIj4KICAgICAgPGRpdiBjbGFzcz0iY2MtaGVhZCI+PGRpdiBjbGFzcz0iY2MtdGl0bGUiPkRheS1FbmQgT3V0Y29tZXM8L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPlN0YXR1cyBhdCBlbmQgb2YgZWFjaCB2aXNpdCBkYXk8L2Rpdj48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY3ciIHN0eWxlPSJoZWlnaHQ6MjIwcHgiPjxjYW52YXMgaWQ9InBvYS1kYXllbmQiPjwvY2FudmFzPjwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJsZWciIGlkPSJwb2EtZGUtbGVnIj48L2Rpdj4KICAgIDwvZGl2PgogIDwvZGl2PgogIDxkaXYgY2xhc3M9ImNjIj4KICAgIDxkaXYgY2xhc3M9ImNjLWhlYWQiPjxkaXYgY2xhc3M9ImNjLXRpdGxlIj5QT0EgRGFpbHkgQnJlYWtkb3duPC9kaXY+PGRpdiBjbGFzcz0iY2Mtc3ViIj5Gb2xsb3ctdXBzIMK3IHdhbGstaW5zIMK3IG91dGNvbWVzIHBlciBkYXRlPC9kaXY+PC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzY3JvbGwteCI+PHRhYmxlIGNsYXNzPSJ3YS10YWJsZSIgaWQ9InBvYS10YmwiPjwvdGFibGU+PC9kaXY+CiAgPC9kaXY+CjwvZGl2PgoKPCEtLSDilZDilZAgS0FNIEZVTk5FTCDilZDilZAgLS0+CjxkaXYgY2xhc3M9InBhbmVsIiBpZD0icC1mdW5uZWwiPgogIDxkaXYgY2xhc3M9ImNjIiBzdHlsZT0ibWFyZ2luLWJvdHRvbToxMnB4Ij4KICAgIDxkaXYgY2xhc3M9ImNjLWhlYWQiPjxkaXYgY2xhc3M9ImNjLXRpdGxlIj5LQU0td2lzZSBGdWxsIFBpcGVsaW5lIEZ1bm5lbDwvZGl2PjxkaXYgY2xhc3M9ImNjLXN1YiI+SW50cm8g4oaSIEZvbGxvdy11cCDihpIgQ29udHJhY3QgU2lnbmluZyDihpIgRG9jIENvbGxlY3Rpb24g4oaSIE9uYm9hcmRlZCDCtyBmcm9tIENSTSBkYXRhIMK3IGZpbHRlcmVkIGRhdGUgcmFuZ2U8L2Rpdj48L2Rpdj4KICAgIDxkaXYgY2xhc3M9InNjcm9sbC14Ij48dGFibGUgY2xhc3M9ImtmLXRhYmxlIiBpZD0ia2YtdGFibGUiPjwvdGFibGU+PC9kaXY+CiAgPC9kaXY+CiAgPGRpdiBjbGFzcz0iZzIiPgogICAgPGRpdiBjbGFzcz0iY2MiPgogICAgICA8ZGl2IGNsYXNzPSJjYy1oZWFkIj48ZGl2IGNsYXNzPSJjYy10aXRsZSI+Q29udmVyc2lvbiBieSBTdGFnZTwvZGl2PjxkaXYgY2xhc3M9ImNjLXN1YiI+VG90YWxzIGFjcm9zcyBhbGwgS0FNcyBpbiByYW5nZTwvZGl2PjwvZGl2PgogICAgICA8ZGl2IGlkPSJmbi1iYXJzIj48L2Rpdj4KICAgIDwvZGl2PgogICAgPGRpdiBjbGFzcz0iY2MiPgogICAgICA8ZGl2IGNsYXNzPSJjYy1oZWFkIj48ZGl2IGNsYXNzPSJjYy10aXRsZSI+T25ib2FyZGluZyBDb252ZXJzaW9uIFJhdGU8L2Rpdj48ZGl2IGNsYXNzPSJjYy1zdWIiPk9uYm9hcmRlZCDDtyBJbnRybyBNZWV0aW5ncyBwZXIgS0FNPC9kaXY+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImN3IiBzdHlsZT0iaGVpZ2h0OjI4MHB4Ij48Y2FudmFzIGlkPSJmbi1jb252Ij48L2NhbnZhcz48L2Rpdj4KICAgIDwvZGl2PgogIDwvZGl2Pgo8L2Rpdj4KCjwhLS0g4pWQ4pWQIFZJU0lUIExPRyDilZDilZAgLS0+CjxkaXYgY2xhc3M9InBhbmVsIiBpZD0icC12aXNpdHMiPgogIDxkaXYgY2xhc3M9Imtyb3cga3JvdzQiIGlkPSJ2bC1rcGlzIj48L2Rpdj4KICA8ZGl2IGNsYXNzPSJnMiIgc3R5bGU9Im1hcmdpbi1ib3R0b206MTJweCI+CiAgICA8ZGl2IGNsYXNzPSJjYyI+CiAgICAgIDxkaXYgY2xhc3M9ImNjLWhlYWQiPjxkaXYgY2xhc3M9ImNjLXRpdGxlIj5EYWlseSBWaXNpdCBBY3Rpdml0eTwvZGl2PjxkaXYgY2xhc3M9ImNjLXN1YiI+V2hhdHNBcHAgdmlzaXQgbG9nIMK3IGVudHJpZXMgcGVyIGRheTwvZGl2PjwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJjdyIgc3R5bGU9ImhlaWdodDoyMDBweCI+PGNhbnZhcyBpZD0idmwtZGFpbHkiPjwvY2FudmFzPjwvZGl2PgogICAgPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJjYyI+CiAgICAgIDxkaXYgY2xhc3M9ImNjLWhlYWQiPjxkaXYgY2xhc3M9ImNjLXRpdGxlIj5WaXNpdHMgYnkgS0FNPC9kaXY+PGRpdiBjbGFzcz0iY2Mtc3ViIj5XaG8gbG9nZ2VkIHZpc2l0cyBpbiBzZWxlY3RlZCByYW5nZTwvZGl2PjwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJjdyIgc3R5bGU9ImhlaWdodDoyMDBweCI+PGNhbnZhcyBpZD0idmwta2FtIj48L2NhbnZhcz48L2Rpdj4KICAgIDwvZGl2PgogIDwvZGl2PgogIDxkaXYgY2xhc3M9ImNjIj4KICAgIDxkaXYgY2xhc3M9ImNjLWhlYWQiPgogICAgICA8ZGl2IGNsYXNzPSJjYy10aXRsZSI+VmlzaXQgTG9nPC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImNjLXN1YiI+V2hhdHNBcHAgZGFpbHkgdGVhbSBsb2cgwrcgY29tbWVudHMgYW5kIGZvbGxvdy11cCBub3RlczwvZGl2PgogICAgPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzY3JvbGwteCBzY3JvbGwteSI+PHRhYmxlIGNsYXNzPSJ3YS10YWJsZSIgaWQ9InZsLXRibCI+PC90YWJsZT48L2Rpdj4KICA8L2Rpdj4KPC9kaXY+Cgo8L2Rpdj48IS0tIC9jb250ZW50IC0tPgo8L2Rpdj48IS0tIC9tYWluIC0tPgoKPHNjcmlwdD4KLy8g4pSA4pSAIERBVEEg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNvbnN0IFJBVyA9IA==").decode('utf-8')
    tmpl_after  = base64.b64decode("OwoKY29uc3QgUEFMPVsnI2Y1YTYyMycsJyMyMmM1NWUnLCcjNGE5MGQ5JywnI2Y0M2Y1ZScsJyM4YjVjZjYnLCcjMTRiOGE2JywnI2Y5NzMxNicsJyMwNmI2ZDQnLCcjZTg3OWY5JywnI2EzZTYzNScsJyNmYWNjMTUnLCcjZmI3MTg1JywnIzM4YmRmOCcsJyNmYmJmMjQnXTsKCmxldCBzdGF0ZSA9IHsgZnJvbTonMjAyNi0wMi0wMScsIHRvOicyMDI2LTAzLTE4Jywga2FtOidhbGwnLCB6b25lOidhbGwnIH07CmxldCBjaGFydHMgPSB7fTsKCi8vIOKUgOKUgCBIRUxQRVJTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApjb25zdCBpblJhbmdlID0gKGR0KSA9PiBkdCAmJiBkdCA+PSBzdGF0ZS5mcm9tICYmIGR0IDw9IHN0YXRlLnRvOwpjb25zdCBieUtBTSA9IChyLGs9J2thbScpID0+IHN0YXRlLmthbT09PSdhbGwnIHx8IHJba109PT1zdGF0ZS5rYW07CmNvbnN0IGJ5S0FNdiA9ICh2KSA9PiBzdGF0ZS5rYW09PT0nYWxsJyB8fCB2PT09c3RhdGUua2FtOwpjb25zdCBmbXQwID0gbiA9PiBNYXRoLnJvdW5kKG58fDApLnRvTG9jYWxlU3RyaW5nKCk7CmNvbnN0IGZtdFAgPSAoYSxiKSA9PiBiID8gKGEvYioxMDApLnRvRml4ZWQoMSkrJyUnIDogJzAlJzsKY29uc3QgZGMgPSBpZCA9PiB7IGlmKGNoYXJ0c1tpZF0pe2NoYXJ0c1tpZF0uZGVzdHJveSgpO2RlbGV0ZSBjaGFydHNbaWRdO30gfTsKCmZ1bmN0aW9uIGthbU5hbWUoZSl7IGlmKCFlfHxlPT09J05vbmUnKSByZXR1cm4gbnVsbDsgaWYoIWUuaW5jbHVkZXMoJ0AnKSkgcmV0dXJuIGU7IHJldHVybiBlLnNwbGl0KCdAJylbMF0uc3BsaXQoJy4nKS5tYXAodz0+d1swXS50b1VwcGVyQ2FzZSgpK3cuc2xpY2UoMSkpLmpvaW4oJyAnKTsgfQoKLy8g4pSA4pSAIElOSVQgRklMVEVSUyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gaW5pdEZpbHRlcnMoKXsKICBjb25zdCBrYW1zID0gWy4uLm5ldyBTZXQoWwogICAgLi4uUkFXLmNybS5tYXAocj0+ci5rYW0pLCAuLi5SQVcucG9hLm1hcChyPT5yLmthbSksCiAgICAuLi5SQVcud2EubWFwKHI9PnIua2FtKSwgLi4uUkFXLmRvY3MubWFwKHI9PnIua2FtKQogIF0pXS5maWx0ZXIoQm9vbGVhbikuc29ydCgpOwogIAogIGNvbnN0IHpvbmVzID0gWy4uLm5ldyBTZXQoWwogICAgLi4uUkFXLnBvYS5tYXAocj0+ci56b25lKSwgLi4uUkFXLmRvY3MubWFwKHI9PnIuem9uZSkKICBdKV0uZmlsdGVyKEJvb2xlYW4pLmZpbHRlcih6PT4hei5pbmNsdWRlcygnbWVudGlvbicpKS5zb3J0KCk7CgogIFsnc2VsLWthbScsJ2RwLWthbSddLmZvckVhY2goaWQ9PnsKICAgIGNvbnN0IGVsID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoaWQpOwogICAgaWYoIWVsKSByZXR1cm47CiAgICBlbC5pbm5lckhUTUwgPSAnPG9wdGlvbiB2YWx1ZT0iYWxsIj5BbGwgS0FNczwvb3B0aW9uPicgKyBrYW1zLm1hcChrPT5gPG9wdGlvbiB2YWx1ZT0iJHtrfSI+JHtrfTwvb3B0aW9uPmApLmpvaW4oJycpOwogIH0pOwogIAogIGNvbnN0IHpzID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ3NlbC16b25lJyk7CiAgenMuaW5uZXJIVE1MID0gJzxvcHRpb24gdmFsdWU9ImFsbCI+QWxsIFpvbmVzPC9vcHRpb24+JyArIHpvbmVzLm1hcCh6PT5gPG9wdGlvbiB2YWx1ZT0iJHt6fSI+JHt6fTwvb3B0aW9uPmApLmpvaW4oJycpOwp9CgovLyDilIDilIAgREFURSBQUkVTRVRTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApmdW5jdGlvbiBzZXRQcmVzZXQocCl7CiAgZG9jdW1lbnQucXVlcnlTZWxlY3RvckFsbCgnLnByLWJ0bicpLmZvckVhY2goYj0+Yi5jbGFzc0xpc3QucmVtb3ZlKCdvbicpKTsKICBldmVudC50YXJnZXQuY2xhc3NMaXN0LmFkZCgnb24nKTsKICBjb25zdCB0b2RheSA9IG5ldyBEYXRlKCkudG9JU09TdHJpbmcoKS5zbGljZSgwLDEwKTsKICBjb25zdCB5ZXN0ID0gbmV3IERhdGUoRGF0ZS5ub3coKS04NjQwMDAwMCkudG9JU09TdHJpbmcoKS5zbGljZSgwLDEwKTsKICBjb25zdCBkNyA9IG5ldyBEYXRlKERhdGUubm93KCktNyo4NjQwMDAwMCkudG9JU09TdHJpbmcoKS5zbGljZSgwLDEwKTsKICBjb25zdCBtMSA9IHRvZGF5LnNsaWNlKDAsNykrJy0wMSc7CiAgaWYocD09PSd0b2RheScpe3N0YXRlLmZyb209c3RhdGUudG89dG9kYXk7fQogIGVsc2UgaWYocD09PSd5ZXN0ZXJkYXknKXtzdGF0ZS5mcm9tPXN0YXRlLnRvPXllc3Q7fQogIGVsc2UgaWYocD09PSc3ZCcpe3N0YXRlLmZyb209ZDc7c3RhdGUudG89dG9kYXk7fQogIGVsc2UgaWYocD09PSdtdGQnKXtzdGF0ZS5mcm9tPW0xO3N0YXRlLnRvPXRvZGF5O30KICBlbHNle3N0YXRlLmZyb209JzIwMjYtMDItMDEnO3N0YXRlLnRvPScyMDI2LTA2LTMwJzt9CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2RyLWZyb20nKS52YWx1ZT1zdGF0ZS5mcm9tOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdkci10bycpLnZhbHVlPXN0YXRlLnRvOwogIGFwcGx5RmlsdGVycygpOwp9CgpmdW5jdGlvbiBhcHBseUZpbHRlcnMoKXsKICBzdGF0ZS5mcm9tID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2RyLWZyb20nKS52YWx1ZTsKICBzdGF0ZS50byAgID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2RyLXRvJykudmFsdWU7CiAgc3RhdGUua2FtICA9IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzZWwta2FtJykudmFsdWU7CiAgc3RhdGUuem9uZSA9IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdzZWwtem9uZScpLnZhbHVlOwogIC8vIFN5bmMgZHAta2FtCiAgY29uc3QgZHBrID0gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2RwLWthbScpOwogIGlmKGRwayAmJiBzdGF0ZS5rYW0hPT0nYWxsJykgZHBrLnZhbHVlPXN0YXRlLmthbTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgndGItc3ViJykudGV4dENvbnRlbnQgPSBgQ0FQUyBEZWxpdmVyeSDCtyAke3N0YXRlLmZyb219IOKGkiAke3N0YXRlLnRvfSR7c3RhdGUua2FtIT09J2FsbCc/JyDCtyAnK3N0YXRlLmthbTonJ31gOwogIHJlUmVuZGVyQWN0aXZlKCk7Cn0KCmxldCBhY3RpdmVUYWIgPSAnb3ZlcnZpZXcnOwpmdW5jdGlvbiBnbyh0YWIpewogIGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5wYW5lbCcpLmZvckVhY2gocD0+cC5jbGFzc0xpc3QucmVtb3ZlKCdvbicpKTsKICBkb2N1bWVudC5xdWVyeVNlbGVjdG9yQWxsKCcuc2ItaXRlbScpLmZvckVhY2goaT0+aS5jbGFzc0xpc3QucmVtb3ZlKCdvbicpKTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgncC0nK3RhYikuY2xhc3NMaXN0LmFkZCgnb24nKTsKICBkb2N1bWVudC5xdWVyeVNlbGVjdG9yQWxsKCcuc2ItaXRlbScpLmZvckVhY2goaT0+e2lmKGkuZ2V0QXR0cmlidXRlKCdvbmNsaWNrJykmJmkuZ2V0QXR0cmlidXRlKCdvbmNsaWNrJykuaW5jbHVkZXMoIiciK3RhYisiJyIpKWkuY2xhc3NMaXN0LmFkZCgnb24nKTt9KTsKICBhY3RpdmVUYWI9dGFiOwogIHJlbmRlclRhYih0YWIpOwp9Cgpjb25zdCByZW5kZXJlZCA9IG5ldyBTZXQoKTsKZnVuY3Rpb24gcmVSZW5kZXJBY3RpdmUoKXsgcmVuZGVyZWQuY2xlYXIoKTsgcmVuZGVyVGFiKGFjdGl2ZVRhYik7IH0KZnVuY3Rpb24gcmVuZGVyVGFiKHRhYil7CiAgaWYocmVuZGVyZWQuaGFzKHRhYikpIHJldHVybjsKICByZW5kZXJlZC5hZGQodGFiKTsKICBjb25zdCBmbj17b3ZlcnZpZXc6cmVuZGVyT3ZlcnZpZXcsbG9jYXRpb246cmVuZGVyTG9jYXRpb24sa2FtOnJlbmRlcktBTSxjb250cmFjdHM6cmVuZGVyQ29udHJhY3RzLGRvY3VtZW50czpyZW5kZXJEb2N1bWVudHMsZGF5cGxhbjpyZW5kZXJEYXlQbGFuLHBvYTpyZW5kZXJQT0EsZnVubmVsOnJlbmRlckZ1bm5lbCx2aXNpdHM6cmVuZGVyVmlzaXRzfTsKICBpZihmblt0YWJdKSBmblt0YWJdKCk7Cn0KCi8vIOKUgOKUgCBGSUxURVJFRCBEQVRBIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApmdW5jdGlvbiBmaWx0ZXJlZE1hc3RlcigpeyByZXR1cm4gUkFXLm1hc3Rlci5maWx0ZXIocj0+aW5SYW5nZShyLmR0KSYmYnlLQU12KHIua2FtKSYmKHN0YXRlLnpvbmU9PT0nYWxsJ3x8ci56b25lPT09c3RhdGUuem9uZSkpOyB9CmZ1bmN0aW9uIGZpbHRlcmVkRG9jcygpeyAgIHJldHVybiBSQVcuZG9jcy5maWx0ZXIocj0+aW5SYW5nZShyLmR0KSYmYnlLQU12KHIua2FtKSYmKHN0YXRlLnpvbmU9PT0nYWxsJ3x8ci56b25lPT09c3RhdGUuem9uZSkpOyB9CmZ1bmN0aW9uIGZpbHRlcmVkQ1JNKCl7ICAgIHJldHVybiBSQVcuY3JtLmZpbHRlcihyPT5pblJhbmdlKHIudmQpJiZieUtBTXYoci5rYW0pKTsgfQpmdW5jdGlvbiBmaWx0ZXJlZFBPQSgpeyAgICByZXR1cm4gUkFXLnBvYS5maWx0ZXIocj0+aW5SYW5nZShyLnZkKSYmYnlLQU12KHIua2FtKSYmKHN0YXRlLnpvbmU9PT0nYWxsJ3x8ci56b25lPT09c3RhdGUuem9uZSkpOyB9CmZ1bmN0aW9uIGZpbHRlcmVkV0EoKXsgICAgIHJldHVybiBSQVcud2EuZmlsdGVyKHI9PmluUmFuZ2Uoci52ZCkmJmJ5S0FNdihyLmthbSkpOyB9CmZ1bmN0aW9uIGZpbHRlcmVkRlUoKXsgICAgIHJldHVybiBSQVcuZnUuZmlsdGVyKHI9PmluUmFuZ2Uoci5mZCkmJmJ5S0FNdihyLmthbSkpOyB9CgovLyDilIDilIAgT1ZFUlZJRVcg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmZ1bmN0aW9uIHJlbmRlck92ZXJ2aWV3KCl7CiAgY29uc3QgbWFzdGVyID0gZmlsdGVyZWRNYXN0ZXIoKSwgZG9jcyA9IGZpbHRlcmVkRG9jcygpLCBjcm0gPSBmaWx0ZXJlZENSTSgpOwogIGNvbnN0IHRvdGFsID0gbWFzdGVyLmxlbmd0aDsKICBjb25zdCBzdWJtaXR0ZWQgPSBtYXN0ZXIuZmlsdGVyKHI9PnIuc3RhdHVzJiZyLnN0YXR1cy50b0xvd2VyQ2FzZSgpLmluY2x1ZGVzKCdzdWJtaXQnKSkubGVuZ3RoOwogIGNvbnN0IHBlbmRpbmcgPSB0b3RhbCAtIHN1Ym1pdHRlZDsKICAKICAvLyBLUElzCiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ292LWtwaXMnKS5pbm5lckhUTUwgPSBbCiAgICB7bGJsOidUb3RhbCBPbmJvYXJkaW5ncycsdmFsOmZtdDAodG90YWwpLHN1YjonSW4gc2VsZWN0ZWQgcGVyaW9kJyxwaWxsOidrYzEnLHB0eHQ6JycscGNsczoncGEnfSwKICAgIHtsYmw6J0NvbnRyYWN0cyBTdWJtaXR0ZWQnLHZhbDpmbXQwKHN1Ym1pdHRlZCksc3ViOmZtdFAoc3VibWl0dGVkLHRvdGFsKSsnIHJhdGUnLHBpbGw6J2tjMicscHR4dDon4oaRIHN1Ym1pdHRlZCcscGNsczoncGcnfSwKICAgIHtsYmw6J0NvbnRyYWN0cyBQZW5kaW5nJyx2YWw6Zm10MChwZW5kaW5nKSxzdWI6J0FjdGlvbiByZXF1aXJlZCcscGlsbDona2MzJyxwdHh0OnBlbmRpbmc+MD8nTmVlZHMgYWN0aW9uJzonQWxsIGNsZWFyJyxwY2xzOnBlbmRpbmc+MD8ncHInOidwZyd9LAogICAge2xibDonQ1JNIFZpc2l0cycsdmFsOmZtdDAoY3JtLmxlbmd0aCksc3ViOidJbiBzZWxlY3RlZCByYW5nZScscGlsbDona2M0JyxwdHh0OidBbGwgbWVldGluZyB0eXBlcycscGNsczoncGInfSwKICAgIHtsYmw6J0ZvbGxvdy11cHMgRHVlJyx2YWw6Zm10MChmaWx0ZXJlZEZVKCkubGVuZ3RoKSxzdWI6J0luIGRhdGUgcmFuZ2UnLHBpbGw6J2tjNScscHR4dDonQ1JNICsgV2hhdHNBcHAnLHBjbHM6J3B2J30sCiAgXS5tYXAoKGssaSk9PmA8ZGl2IGNsYXNzPSJrY2FyZCBrYyR7aSsxfSI+PGRpdiBjbGFzcz0ia2xibCI+JHtrLmxibH08L2Rpdj48ZGl2IGNsYXNzPSJrdmFsIj4ke2sudmFsfTwvZGl2PjxkaXYgY2xhc3M9ImtzdWIiPiR7ay5zdWJ9PC9kaXY+PHNwYW4gY2xhc3M9ImtwaWxsICR7ay5wY2xzfSI+JHtrLnB0eHR9PC9zcGFuPjwvZGl2PmApLmpvaW4oJycpOwoKICAvLyBab25lIGJhcnMKICBjb25zdCB6b25lTWFwPXt9OwogIGRvY3MuZm9yRWFjaChyPT57Y29uc3Qgej1yLnpvbmV8fCdVbmtub3duJztpZighei5pbmNsdWRlcygnbWVudGlvbicpKXpvbmVNYXBbel09KHpvbmVNYXBbel18fDApKzE7fSk7CiAgY29uc3QgelNvcnRlZD1PYmplY3QuZW50cmllcyh6b25lTWFwKS5zb3J0KChhLGIpPT5iWzFdLWFbMV0pOwogIGNvbnN0IG1heFo9TWF0aC5tYXgoLi4uelNvcnRlZC5tYXAoej0+elsxXSksMSk7CiAgY29uc3QgekNvbG9ycz1bJyNmNWE2MjMnLCcjNGE5MGQ5JywnIzIyYzU1ZScsJyNmNDNmNWUnLCcjOGI1Y2Y2JywnIzE0YjhhNiddOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdvdi16b25lLWJhcnMnKS5pbm5lckhUTUwgPSB6U29ydGVkLm1hcCgoW25hbWUsdmFsXSxpKT0+CiAgICBgPGRpdiBjbGFzcz0iaGJhci1pdGVtIj48ZGl2IGNsYXNzPSJoYmFyLW5hbWUiIHN0eWxlPSJ3aWR0aDoxNjBweCI+JHtuYW1lfTwvZGl2PjxkaXYgY2xhc3M9ImhiYXItdHJhY2siPjxkaXYgY2xhc3M9ImhiYXItZmlsbCIgc3R5bGU9IndpZHRoOiR7TWF0aC5yb3VuZCh2YWwvbWF4WioxMDApfSU7YmFja2dyb3VuZDoke3pDb2xvcnNbaSV6Q29sb3JzLmxlbmd0aF19Ij4ke3ZhbH08L2Rpdj48L2Rpdj48ZGl2IGNsYXNzPSJoYmFyLW51bSI+JHt2YWx9PC9kaXY+PC9kaXY+YAogICkuam9pbignJyk7CgogIC8vIENvbnRyYWN0IGRvbnV0CiAgY29uc3QgY3M9e1N1Ym1pdHRlZDpzdWJtaXR0ZWQsUGVuZGluZzpwZW5kaW5nfTsKICBkYygnb3YtY3QnKTsKICBjaGFydHNbJ292LWN0J109bmV3IENoYXJ0KGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdvdi1jb250cmFjdCcpLHsKICAgIHR5cGU6J2RvdWdobnV0JywKICAgIGRhdGE6e2xhYmVsczpPYmplY3Qua2V5cyhjcyksZGF0YXNldHM6W3tkYXRhOk9iamVjdC52YWx1ZXMoY3MpLGJhY2tncm91bmRDb2xvcjpbJyMyMmM1NWVjYycsJyNmNDNmNWVjYyddLGJvcmRlckNvbG9yOicjMWEyMDMwJyxib3JkZXJXaWR0aDozLGhvdmVyT2Zmc2V0OjZ9XX0sCiAgICBvcHRpb25zOntyZXNwb25zaXZlOnRydWUsbWFpbnRhaW5Bc3BlY3RSYXRpbzpmYWxzZSxjdXRvdXQ6JzY1JScscGx1Z2luczp7bGVnZW5kOntkaXNwbGF5OmZhbHNlfSx0b29sdGlwOntjYWxsYmFja3M6e2xhYmVsOmN0eD0+YCR7Y3R4LmxhYmVsfTogJHtjdHgucmF3fSAoJHtmbXRQKGN0eC5yYXcsdG90YWwpfSlgfX19fQogIH0pOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdvdi1jb250cmFjdC1sZWcnKS5pbm5lckhUTUw9W1snU3VibWl0dGVkJywnIzIyYzU1ZSddLFsnUGVuZGluZycsJyNmNDNmNWUnXV0ubWFwKChbbCxjXSk9PmA8c3BhbiBjbGFzcz0ibGkiPjxzcGFuIGNsYXNzPSJsc3EiIHN0eWxlPSJiYWNrZ3JvdW5kOiR7Y30iPjwvc3Bhbj4ke2x9PC9zcGFuPmApLmpvaW4oJycpOwoKICAvLyBGdW5uZWwgbWluaQogIGNvbnN0IHN0YWdlcz1bWydJbnRybyBNZWV0aW5nJyxjcm0uZmlsdGVyKHI9PnIubXQ9PT0nSW50cm8gTWVldGluZycpLmxlbmd0aCwnIzRhOTBkOSddLAogICAgWydGb2xsb3ctdXAnLGNybS5maWx0ZXIocj0+ci5tdD09PSdGb2xsb3cgdXAnKS5sZW5ndGgsJyNmNWE2MjMnXSwKICAgIFsnT25ib2FyZGVkJyxjcm0uZmlsdGVyKHI9PnIub2ImJnIub2IudG9Mb3dlckNhc2UoKT09PSd5ZXMnKS5sZW5ndGgsJyMyMmM1NWUnXSwKICAgIFsnRm9sbG93LXVwIFJlcXVpcmVkJyxjcm0uZmlsdGVyKHI9PnIuZnUmJnIuZnUudG9Mb3dlckNhc2UoKT09PSd5ZXMnKS5sZW5ndGgsJyM4YjVjZjYnXV07CiAgY29uc3QgbWF4Rj1NYXRoLm1heCguLi5zdGFnZXMubWFwKHM9PnNbMV0pLDEpOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdvdi1mdW5uZWwnKS5pbm5lckhUTUw9c3RhZ2VzLm1hcCgoW2wsdixjXSk9PgogICAgYDxkaXYgY2xhc3M9ImZuLXJvdyI+PGRpdiBjbGFzcz0iZm4tbGFiZWwiPiR7bH08L2Rpdj48ZGl2IGNsYXNzPSJmbi10cmFjayI+PGRpdiBjbGFzcz0iZm4tZmlsbCIgc3R5bGU9IndpZHRoOiR7TWF0aC5yb3VuZCh2L21heEYqMTAwKX0lO2JhY2tncm91bmQ6JHtjfSI+JHt2fTwvZGl2PjwvZGl2PjxkaXYgY2xhc3M9ImZuLWNvdW50Ij4ke3YudG9Mb2NhbGVTdHJpbmcoKX08L2Rpdj48L2Rpdj5gCiAgKS5qb2luKCcnKTsKCiAgLy8gRG9jIG92ZXJ2aWV3IG1pbmkKICBjb25zdCBET0NfRklFTERTPVsnZnNzYWknLCdnc3QnLCdwYW4nLCdjaGVxdWUnLCdtZW51JywnbGFib3VyJ107CiAgY29uc3QgZG9jQ29sb3JzPVsnIzIyYzU1ZScsJyNmNDNmNWUnLCcjNGE5MGQ5JywnI2Y1YTYyMycsJyM4YjVjZjYnLCcjMTRiOGE2J107CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ292LWRvY3MnKS5pbm5lckhUTUw9RE9DX0ZJRUxEUy5tYXAoKGYsaSk9PnsKICAgIGNvbnN0IHllcz1kb2NzLmZpbHRlcihyPT5yW2ZdPT09J3llcycpLmxlbmd0aDsKICAgIGNvbnN0IG5vPWRvY3MuZmlsdGVyKHI9PnJbZl09PT0nbm8nKS5sZW5ndGg7CiAgICBjb25zdCBwY3Q9ZG9jcy5sZW5ndGg/TWF0aC5yb3VuZCh5ZXMvZG9jcy5sZW5ndGgqMTAwKTowOwogICAgcmV0dXJuIGA8ZGl2IGNsYXNzPSJkb2Mtcm93Ij48ZGl2IGNsYXNzPSJkb2MtaGQiPjxzcGFuIGNsYXNzPSJkb2Mtbm0iPiR7Zi50b1VwcGVyQ2FzZSgpfTwvc3Bhbj48c3BhbiBjbGFzcz0iZG9jLXBjdCI+JHtwY3R9JTwvc3Bhbj48L2Rpdj48ZGl2IHN0eWxlPSJkaXNwbGF5OmZsZXg7anVzdGlmeS1jb250ZW50OnNwYWNlLWJldHdlZW47Zm9udC1zaXplOi42NHJlbTtjb2xvcjp2YXIoLS1tdSk7bWFyZ2luLWJvdHRvbTo0cHgiPjxzcGFuIHN0eWxlPSJjb2xvcjp2YXIoLS1nbikiPiR7eWVzfSBjb2xsZWN0ZWQ8L3NwYW4+PHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXJzKSI+JHtub30gcGVuZGluZzwvc3Bhbj48L2Rpdj48ZGl2IGNsYXNzPSJkb2MtYmFyIj48ZGl2IGNsYXNzPSJkb2MtZmlsbCIgc3R5bGU9IndpZHRoOiR7cGN0fSU7YmFja2dyb3VuZDoke2RvY0NvbG9yc1tpXX0iPjwvZGl2PjwvZGl2PjwvZGl2PmA7CiAgfSkuam9pbignJyk7Cn0KCi8vIOKUgOKUgCBMT0NBVElPTiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gcmVuZGVyTG9jYXRpb24oKXsKICBjb25zdCBkb2NzPWZpbHRlcmVkRG9jcygpLCBwb2E9ZmlsdGVyZWRQT0EoKTsKCiAgLy8gWm9uZSBjaGFydAogIGNvbnN0IHptPXt9O2RvY3MuZm9yRWFjaChyPT57Y29uc3Qgej1yLnpvbmV8fCc/JztpZighei5pbmNsdWRlcygnbWVudGlvbicpKXptW3pdPSh6bVt6XXx8MCkrMTt9KTsKICBjb25zdCB6bD1PYmplY3Qua2V5cyh6bSksenY9T2JqZWN0LnZhbHVlcyh6bSk7CiAgZGMoJ2xvYy16Jyk7CiAgY2hhcnRzWydsb2MteiddPW5ldyBDaGFydChkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnbG9jLXpvbmUnKSx7dHlwZTonYmFyJyxkYXRhOntsYWJlbHM6emwsZGF0YXNldHM6W3tkYXRhOnp2LGJhY2tncm91bmRDb2xvcjpQQUwubWFwKGM9PmMrJ2JiJyksYm9yZGVyQ29sb3I6UEFMLGJvcmRlcldpZHRoOjEuNSxib3JkZXJSYWRpdXM6NX1dfSxvcHRpb25zOntyZXNwb25zaXZlOnRydWUsbWFpbnRhaW5Bc3BlY3RSYXRpbzpmYWxzZSxwbHVnaW5zOntsZWdlbmQ6e2Rpc3BsYXk6ZmFsc2V9fSxzY2FsZXM6e3g6e2dyaWQ6e2Rpc3BsYXk6ZmFsc2V9LHRpY2tzOntjb2xvcjonIzdhODdhOCcsZm9udDp7c2l6ZToxMH0sbWF4Um90YXRpb246MzB9fSx5OntncmlkOntjb2xvcjonIzI0MmQ0Mid9LHRpY2tzOntjb2xvcjonIzdhODdhOCcsZm9udDp7c2l6ZToxMH19fX19fSk7CgogIC8vIFBPQSB6b25lCiAgY29uc3QgcHptPXt9O3BvYS5mb3JFYWNoKHI9Pntjb25zdCB6PXIuem9uZXx8Jz8nO2lmKHomJiF6LmluY2x1ZGVzKCdtZW50aW9uJykpcHptW3pdPShwem1bel18fDApKzE7fSk7CiAgY29uc3QgcHpsPU9iamVjdC5rZXlzKHB6bSkscHp2PU9iamVjdC52YWx1ZXMocHptKTsKICBkYygnbG9jLXB6Jyk7CiAgY2hhcnRzWydsb2MtcHonXT1uZXcgQ2hhcnQoZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2xvYy1wb2Etem9uZScpLHt0eXBlOidiYXInLGRhdGE6e2xhYmVsczpwemwsZGF0YXNldHM6W3tkYXRhOnB6dixiYWNrZ3JvdW5kQ29sb3I6JyM0YTkwZDk2NicsYm9yZGVyQ29sb3I6JyM0YTkwZDknLGJvcmRlcldpZHRoOjEuNSxib3JkZXJSYWRpdXM6NX1dfSxvcHRpb25zOntyZXNwb25zaXZlOnRydWUsbWFpbnRhaW5Bc3BlY3RSYXRpbzpmYWxzZSxwbHVnaW5zOntsZWdlbmQ6e2Rpc3BsYXk6ZmFsc2V9fSxzY2FsZXM6e3g6e2dyaWQ6e2Rpc3BsYXk6ZmFsc2V9LHRpY2tzOntjb2xvcjonIzdhODdhOCcsZm9udDp7c2l6ZToxMH0sbWF4Um90YXRpb246MzB9fSx5OntncmlkOntjb2xvcjonIzI0MmQ0Mid9LHRpY2tzOntjb2xvcjonIzdhODdhOCcsZm9udDp7c2l6ZToxMH19fX19fSk7CgogIC8vIExvY2F0aW9uIGJhcgogIGNvbnN0IGxtPXt9O2RvY3MuZm9yRWFjaChyPT57Y29uc3QgbD0oci5sb2N8fCcnKS50cmltKCkudG9Mb3dlckNhc2UoKS5yZXBsYWNlKC9eXHcvLGM9PmMudG9VcHBlckNhc2UoKSk7aWYobCYmbCE9PSdOb25lJyYmIWwuaW5jbHVkZXMoJ21lbnRpb24nKSlsbVtsXT0obG1bbF18fDApKzE7fSk7CiAgY29uc3QgbFNvcnRlZD1PYmplY3QuZW50cmllcyhsbSkuc29ydCgoYSxiKT0+YlsxXS1hWzFdKS5zbGljZSgwLDE4KTsKICBkYygnbG9jLWwnKTsKICBjaGFydHNbJ2xvYy1sJ109bmV3IENoYXJ0KGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdsb2MtbG9jcycpLHt0eXBlOidiYXInLGRhdGE6e2xhYmVsczpsU29ydGVkLm1hcChsPT5sWzBdKSxkYXRhc2V0czpbe2RhdGE6bFNvcnRlZC5tYXAobD0+bFsxXSksYmFja2dyb3VuZENvbG9yOicjNGE5MGQ5NDQnLGJvcmRlckNvbG9yOicjNGE5MGQ5Jyxib3JkZXJXaWR0aDoxLjUsYm9yZGVyUmFkaXVzOjR9XX0sb3B0aW9uczp7aW5kZXhBeGlzOid5JyxyZXNwb25zaXZlOnRydWUsbWFpbnRhaW5Bc3BlY3RSYXRpbzpmYWxzZSxwbHVnaW5zOntsZWdlbmQ6e2Rpc3BsYXk6ZmFsc2V9fSxzY2FsZXM6e3g6e2dyaWQ6e2NvbG9yOicjMjQyZDQyJ30sdGlja3M6e2NvbG9yOicjN2E4N2E4Jyxmb250OntzaXplOjEwfX19LHk6e2dyaWQ6e2Rpc3BsYXk6ZmFsc2V9LHRpY2tzOntjb2xvcjonI2RkZTNmMicsZm9udDp7c2l6ZToxMH19fX19fSk7Cn0KCi8vIOKUgOKUgCBLQU0g4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmZ1bmN0aW9uIHJlbmRlcktBTSgpewogIGNvbnN0IGRvY3M9ZmlsdGVyZWREb2NzKCksIG1hc3Rlcj1maWx0ZXJlZE1hc3RlcigpOwogIGNvbnN0IGttPXt9O2RvY3MuZm9yRWFjaChyPT57aWYoci5rYW0pa21bci5rYW1dPShrbVtyLmthbV18fDApKzE7fSk7CiAgY29uc3Qga1NvcnRlZD1PYmplY3QuZW50cmllcyhrbSkuc29ydCgoYSxiKT0+YlsxXS1hWzFdKTsKICBjb25zdCB0b3RhbD1rU29ydGVkLnJlZHVjZSgocyxrKT0+cytrWzFdLDApOwoKICBkYygna2FtLWInKTsKICBjaGFydHNbJ2thbS1iJ109bmV3IENoYXJ0KGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdrYW0tYmFyJykse3R5cGU6J2JhcicsZGF0YTp7bGFiZWxzOmtTb3J0ZWQubWFwKGs9PmtbMF0pLGRhdGFzZXRzOlt7ZGF0YTprU29ydGVkLm1hcChrPT5rWzFdKSxiYWNrZ3JvdW5kQ29sb3I6UEFMLm1hcChjPT5jKydiYicpLGJvcmRlckNvbG9yOlBBTCxib3JkZXJXaWR0aDoxLjUsYm9yZGVyUmFkaXVzOjV9XX0sb3B0aW9uczp7aW5kZXhBeGlzOid5JyxyZXNwb25zaXZlOnRydWUsbWFpbnRhaW5Bc3BlY3RSYXRpbzpmYWxzZSxwbHVnaW5zOntsZWdlbmQ6e2Rpc3BsYXk6ZmFsc2V9fSxzY2FsZXM6e3g6e2dyaWQ6e2NvbG9yOicjMjQyZDQyJ30sdGlja3M6e2NvbG9yOicjN2E4N2E4Jyxmb250OntzaXplOjEwfX19LHk6e2dyaWQ6e2Rpc3BsYXk6ZmFsc2V9LHRpY2tzOntjb2xvcjonI2RkZTNmMicsZm9udDp7c2l6ZToxMH19fX19fSk7CgogIGRjKCdrYW0tZCcpOwogIGNoYXJ0c1sna2FtLWQnXT1uZXcgQ2hhcnQoZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2thbS1kb251dCcpLHt0eXBlOidkb3VnaG51dCcsZGF0YTp7bGFiZWxzOmtTb3J0ZWQubWFwKGs9PmtbMF0pLGRhdGFzZXRzOlt7ZGF0YTprU29ydGVkLm1hcChrPT5rWzFdKSxiYWNrZ3JvdW5kQ29sb3I6UEFMLm1hcChjPT5jKydjYycpLGJvcmRlckNvbG9yOicjMWEyMDMwJyxib3JkZXJXaWR0aDoyLGhvdmVyT2Zmc2V0OjV9XX0sb3B0aW9uczp7cmVzcG9uc2l2ZTp0cnVlLG1haW50YWluQXNwZWN0UmF0aW86ZmFsc2UsY3V0b3V0Oic1NSUnLHBsdWdpbnM6e2xlZ2VuZDp7ZGlzcGxheTpmYWxzZX19fX0pOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdrYW0tZGxlZycpLmlubmVySFRNTD1rU29ydGVkLm1hcCgoayxpKT0+YDxzcGFuIGNsYXNzPSJsaSI+PHNwYW4gY2xhc3M9ImxzcSIgc3R5bGU9ImJhY2tncm91bmQ6JHtQQUxbaSVQQUwubGVuZ3RoXX0iPjwvc3Bhbj4ke2tbMF0uc3BsaXQoJyAnKVswXX08L3NwYW4+YCkuam9pbignJyk7CgogIC8vIExlYWRlcmJvYXJkCiAgY29uc3QgY29udHJhY3RCeUtBTT17fTttYXN0ZXIuZm9yRWFjaChyPT57aWYoci5rYW0pe2lmKCFjb250cmFjdEJ5S0FNW3Iua2FtXSljb250cmFjdEJ5S0FNW3Iua2FtXT17c3ViOjAsdG90YWw6MH07Y29udHJhY3RCeUtBTVtyLmthbV0udG90YWwrKztpZihyLnN0YXR1cyYmci5zdGF0dXMudG9Mb3dlckNhc2UoKS5pbmNsdWRlcygnc3VibWl0JykpY29udHJhY3RCeUtBTVtyLmthbV0uc3ViKys7fX0pOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdrYW0tbGInKS5pbm5lckhUTUw9YDx0aGVhZD48dHI+PHRoPlJhbms8L3RoPjx0aD5LQU08L3RoPjx0aD5PbmJvYXJkaW5nczwvdGg+PHRoPlNoYXJlPC90aD48dGg+Q29udHJhY3RzIFN1Ym1pdHRlZDwvdGg+PHRoPlByb2dyZXNzPC90aD48L3RyPjwvdGhlYWQ+PHRib2R5PmArCiAgICBrU29ydGVkLm1hcCgoW2ssdl0saSk9PnsKICAgICAgY29uc3QgYz1jb250cmFjdEJ5S0FNW2tdfHx7c3ViOjAsdG90YWw6MH07CiAgICAgIGNvbnN0IHBjdD0odi90b3RhbCoxMDApLnRvRml4ZWQoMSk7CiAgICAgIGNvbnN0IHJjbHM9aT09PTA/J3JrMSc6aT09PTE/J3JrMic6aT09PTI/J3JrMyc6J3Jrbic7CiAgICAgIHJldHVybiBgPHRyPjx0ZD48c3BhbiBjbGFzcz0icmFuay1iYWRnZSAke3JjbHN9Ij4ke2krMX08L3NwYW4+PC90ZD48dGQgc3R5bGU9ImZvbnQtd2VpZ2h0OjYwMCI+JHtrfTwvdGQ+PHRkIHN0eWxlPSJjb2xvcjp2YXIoLS1hbSk7Zm9udC13ZWlnaHQ6NzAwIj4ke3Z9PC90ZD48dGQ+JHtwY3R9JTwvdGQ+PHRkPiR7Yy5zdWJ9LyR7Yy50b3RhbH08L3RkPjx0ZCBzdHlsZT0ibWluLXdpZHRoOjEwMHB4Ij48ZGl2IGNsYXNzPSJwcm9nLXdyYXAiPjxkaXYgY2xhc3M9InByb2ctZmlsbCIgc3R5bGU9IndpZHRoOiR7cGN0fSU7YmFja2dyb3VuZDoke1BBTFtpJVBBTC5sZW5ndGhdfSI+PC9kaXY+PC9kaXY+PC90ZD48L3RyPmA7CiAgICB9KS5qb2luKCcnKStgPC90Ym9keT5gOwp9CgovLyDilIDilIAgQ09OVFJBQ1RTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApmdW5jdGlvbiByZW5kZXJDb250cmFjdHMoKXsKICBjb25zdCBtYXN0ZXI9ZmlsdGVyZWRNYXN0ZXIoKTsKICBjb25zdCBjcz17U3VibWl0dGVkOjAsJ05vdCBTdWJtaXR0ZWQnOjAsRHJvcHBlZDowLCdNaXNzaW5nL0NhbmNlbGxlZCc6MH07CiAgbWFzdGVyLmZvckVhY2gocj0+ewogICAgY29uc3Qgcz0oci5zdGF0dXN8fCcnKS50b0xvd2VyQ2FzZSgpOwogICAgaWYocy5pbmNsdWRlcygnc3VibWl0JykmJiFzLmluY2x1ZGVzKCdub3QnKSljcy5TdWJtaXR0ZWQrKzsKICAgIGVsc2UgaWYocy5pbmNsdWRlcygnbm90Jyl8fHMuaW5jbHVkZXMoJ3Vua25vd24nKXx8IXMpY3NbJ05vdCBTdWJtaXR0ZWQnXSsrOwogICAgZWxzZSBpZihzLmluY2x1ZGVzKCdkcm9wJykpY3MuRHJvcHBlZCsrOwogICAgZWxzZSBjc1snTWlzc2luZy9DYW5jZWxsZWQnXSsrOwogIH0pOwogIGNvbnN0IHRvdGFsPW1hc3Rlci5sZW5ndGg7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2N0LWtwaXMnKS5pbm5lckhUTUw9WwogICAge2w6J1N1Ym1pdHRlZCcsdjpjcy5TdWJtaXR0ZWQscGN0OmZtdFAoY3MuU3VibWl0dGVkLHRvdGFsKSxwY2xzOidwZycsa2M6J2tjMSd9LAogICAge2w6J05vdCBTdWJtaXR0ZWQnLHY6Y3NbJ05vdCBTdWJtaXR0ZWQnXSxwY3Q6Zm10UChjc1snTm90IFN1Ym1pdHRlZCddLHRvdGFsKSxwY2xzOidwcicsa2M6J2tjMid9LAogICAge2w6J0Ryb3BwZWQnLHY6Y3MuRHJvcHBlZCxwY3Q6Zm10UChjcy5Ecm9wcGVkLHRvdGFsKSxwY2xzOidwYScsa2M6J2tjMyd9LAogICAge2w6J1RvdGFsJyx2OnRvdGFsLHBjdDonMTAwJScscGNsczoncGInLGtjOidrYzQnfSwKICBdLm1hcChrPT5gPGRpdiBjbGFzcz0ia2NhcmQgJHtrLmtjfSI+PGRpdiBjbGFzcz0ia2xibCI+JHtrLmx9PC9kaXY+PGRpdiBjbGFzcz0ia3ZhbCI+JHtrLnZ9PC9kaXY+PHNwYW4gY2xhc3M9ImtwaWxsICR7ay5wY2xzfSI+JHtrLnBjdH08L3NwYW4+PC9kaXY+YCkuam9pbignJyk7CgogIGNvbnN0IGNDb2xvcnM9WycjMjJjNTVlJywnI2Y0M2Y1ZScsJyNmNWE2MjMnLCcjN2E4N2E4J107CiAgZGMoJ2N0LWInKTsKICBjaGFydHNbJ2N0LWInXT1uZXcgQ2hhcnQoZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2N0LWJhcicpLHt0eXBlOidiYXInLGRhdGE6e2xhYmVsczpPYmplY3Qua2V5cyhjcyksZGF0YXNldHM6W3tkYXRhOk9iamVjdC52YWx1ZXMoY3MpLGJhY2tncm91bmRDb2xvcjpjQ29sb3JzLm1hcChjPT5jKydiYicpLGJvcmRlckNvbG9yOmNDb2xvcnMsYm9yZGVyV2lkdGg6MS41LGJvcmRlclJhZGl1czo2fV19LG9wdGlvbnM6e3Jlc3BvbnNpdmU6dHJ1ZSxtYWludGFpbkFzcGVjdFJhdGlvOmZhbHNlLHBsdWdpbnM6e2xlZ2VuZDp7ZGlzcGxheTpmYWxzZX19LHNjYWxlczp7eDp7Z3JpZDp7ZGlzcGxheTpmYWxzZX0sdGlja3M6e2NvbG9yOicjN2E4N2E4Jyxmb250OntzaXplOjExfX19LHk6e2dyaWQ6e2NvbG9yOicjMjQyZDQyJ30sdGlja3M6e2NvbG9yOicjN2E4N2E4Jyxmb250OntzaXplOjEwfX19fX19KTsKCiAgLy8gS0FNLXdpc2UgY29udHJhY3QgcmF0ZQogIGNvbnN0IGttPXt9OwogIG1hc3Rlci5mb3JFYWNoKHI9PntpZighci5rYW0pcmV0dXJuO2lmKCFrbVtyLmthbV0pa21bci5rYW1dPXtzOjAsdDowfTtrbVtyLmthbV0udCsrO2lmKChyLnN0YXR1c3x8JycpLnRvTG93ZXJDYXNlKCkuaW5jbHVkZXMoJ3N1Ym1pdCcpJiYhKHIuc3RhdHVzfHwnJykudG9Mb3dlckNhc2UoKS5pbmNsdWRlcygnbm90Jykpa21bci5rYW1dLnMrKzt9KTsKICBjb25zdCBrbD1PYmplY3Qua2V5cyhrbSkuc29ydCgoYSxiKT0+a21bYl0ucy9rbVtiXS50LWttW2FdLnMva21bYV0udCk7CiAgZGMoJ2N0LWttJyk7CiAgY2hhcnRzWydjdC1rbSddPW5ldyBDaGFydChkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnY3Qta2FtJykse3R5cGU6J2JhcicsZGF0YTp7bGFiZWxzOmtsLGRhdGFzZXRzOlt7bGFiZWw6J1N1Ym1pdHRlZCcsZGF0YTprbC5tYXAoaz0+a21ba10ucyksYmFja2dyb3VuZENvbG9yOicjMjJjNTVlY2MnLGJvcmRlckNvbG9yOicjMjJjNTVlJyxib3JkZXJXaWR0aDoxLjUsYm9yZGVyUmFkaXVzOjR9LHtsYWJlbDonUGVuZGluZycsZGF0YTprbC5tYXAoaz0+a21ba10udC1rbVtrXS5zKSxiYWNrZ3JvdW5kQ29sb3I6JyNmNDNmNWVjYycsYm9yZGVyQ29sb3I6JyNmNDNmNWUnLGJvcmRlcldpZHRoOjEuNSxib3JkZXJSYWRpdXM6NH1dfSxvcHRpb25zOntyZXNwb25zaXZlOnRydWUsbWFpbnRhaW5Bc3BlY3RSYXRpbzpmYWxzZSxwbHVnaW5zOntsZWdlbmQ6e2Rpc3BsYXk6ZmFsc2V9fSxzY2FsZXM6e3g6e3N0YWNrZWQ6dHJ1ZSxncmlkOntkaXNwbGF5OmZhbHNlfSx0aWNrczp7Y29sb3I6JyM3YTg3YTgnLGZvbnQ6e3NpemU6OX0sbWF4Um90YXRpb246NDB9fSx5OntzdGFja2VkOnRydWUsZ3JpZDp7Y29sb3I6JyMyNDJkNDInfSx0aWNrczp7Y29sb3I6JyM3YTg3YTgnLGZvbnQ6e3NpemU6MTB9fX19fX0pOwp9CgovLyDilIDilIAgRE9DVU1FTlRTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApmdW5jdGlvbiByZW5kZXJEb2N1bWVudHMoKXsKICBjb25zdCBkb2NzPWZpbHRlcmVkRG9jcygpOwogIGNvbnN0IERPQ19GSUVMRFM9Wydmc3NhaScsJ2dzdCcsJ3BhbicsJ2NoZXF1ZScsJ21lbnUnLCdsYWJvdXInXTsKICBjb25zdCBkU3RhdHM9RE9DX0ZJRUxEUy5tYXAoZj0+KHtuYW1lOmYudG9VcHBlckNhc2UoKSx5ZXM6ZG9jcy5maWx0ZXIocj0+cltmXT09PSd5ZXMnKS5sZW5ndGgsbm86ZG9jcy5maWx0ZXIocj0+cltmXT09PSdubycpLmxlbmd0aCx0b3RhbDpkb2NzLmxlbmd0aH0pKTsKCiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2RvYy1rcGlzJykuaW5uZXJIVE1MPWRTdGF0cy5tYXAoKGQsaSk9PnsKICAgIGNvbnN0IHBjdD1kLnRvdGFsP01hdGgucm91bmQoZC55ZXMvZC50b3RhbCoxMDApOjA7CiAgICBjb25zdCBwY2xzPXBjdD43MD8ncGcnOnBjdD40MD8ncGEnOidwcic7CiAgICByZXR1cm4gYDxkaXYgY2xhc3M9ImtjYXJkIGtjJHsoaSU1KSsxfSI+PGRpdiBjbGFzcz0ia2xibCI+JHtkLm5hbWV9PC9kaXY+PGRpdiBjbGFzcz0ia3ZhbCI+JHtkLnllc308L2Rpdj48ZGl2IGNsYXNzPSJrc3ViIj4ke3BjdH0lIGNvbGxlY3RlZDwvZGl2PjxzcGFuIGNsYXNzPSJrcGlsbCAke3BjbHN9Ij4ke2Qubm99IHBlbmRpbmc8L3NwYW4+PC9kaXY+YDsKICB9KS5qb2luKCcnKTsKCiAgZGMoJ2RvYy1iJyk7CiAgY2hhcnRzWydkb2MtYiddPW5ldyBDaGFydChkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnZG9jLWJhcicpLHt0eXBlOidiYXInLGRhdGE6e2xhYmVsczpET0NfRklFTERTLm1hcChmPT5mLnRvVXBwZXJDYXNlKCkpLGRhdGFzZXRzOlt7bGFiZWw6J0NvbGxlY3RlZCcsZGF0YTpkU3RhdHMubWFwKGQ9PmQueWVzKSxiYWNrZ3JvdW5kQ29sb3I6JyMyMmM1NWVjYycsYm9yZGVyQ29sb3I6JyMyMmM1NWUnLGJvcmRlcldpZHRoOjEuNSxib3JkZXJSYWRpdXM6NX0se2xhYmVsOidQZW5kaW5nJyxkYXRhOmRTdGF0cy5tYXAoZD0+ZC5ubyksYmFja2dyb3VuZENvbG9yOicjZjQzZjVlY2MnLGJvcmRlckNvbG9yOicjZjQzZjVlJyxib3JkZXJXaWR0aDoxLjUsYm9yZGVyUmFkaXVzOjV9XX0sb3B0aW9uczp7cmVzcG9uc2l2ZTp0cnVlLG1haW50YWluQXNwZWN0UmF0aW86ZmFsc2UscGx1Z2luczp7bGVnZW5kOntkaXNwbGF5OmZhbHNlfX0sc2NhbGVzOnt4OntncmlkOntkaXNwbGF5OmZhbHNlfSx0aWNrczp7Y29sb3I6JyM3YTg3YTgnLGZvbnQ6e3NpemU6MTJ9fX0seTp7Z3JpZDp7Y29sb3I6JyMyNDJkNDInfSx0aWNrczp7Y29sb3I6JyM3YTg3YTgnLGZvbnQ6e3NpemU6MTB9fX19fX0pOwoKICBjb25zdCBkb2NDb2xzPVsnIzIyYzU1ZScsJyNmNDNmNWUnLCcjNGE5MGQ5JywnI2Y1YTYyMycsJyM4YjVjZjYnLCcjMTRiOGE2J107CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2RvYy1kZXRhaWwnKS5pbm5lckhUTUw9ZFN0YXRzLm1hcCgoZCxpKT0+ewogICAgY29uc3QgcGN0PWQudG90YWw/TWF0aC5yb3VuZChkLnllcy9kLnRvdGFsKjEwMCk6MDsKICAgIHJldHVybiBgPGRpdiBjbGFzcz0iZG9jLXJvdyI+PGRpdiBjbGFzcz0iZG9jLWhkIj48c3BhbiBjbGFzcz0iZG9jLW5tIj4ke2QubmFtZX0g4oCUICR7cGN0fSUgY29sbGVjdGVkPC9zcGFuPjxzcGFuIGNsYXNzPSJkb2MtcGN0Ij4ke2QueWVzfS8ke2QudG90YWx9PC9zcGFuPjwvZGl2PjxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDtqdXN0aWZ5LWNvbnRlbnQ6c3BhY2UtYmV0d2Vlbjtmb250LXNpemU6LjYzcmVtO2NvbG9yOnZhcigtLW11KTttYXJnaW4tYm90dG9tOjVweCI+PHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLWduKSI+4pyTICR7ZC55ZXN9IGNvbGxlY3RlZDwvc3Bhbj48c3BhbiBzdHlsZT0iY29sb3I6dmFyKC0tcnMpIj7inJcgJHtkLm5vfSBwZW5kaW5nPC9zcGFuPjxzcGFuPiR7ZC50b3RhbC1kLnllcy1kLm5vfSBubyBkYXRhPC9zcGFuPjwvZGl2PjxkaXYgY2xhc3M9ImRvYy1iYXIiPjxkaXYgY2xhc3M9ImRvYy1maWxsIiBzdHlsZT0id2lkdGg6JHtwY3R9JTtiYWNrZ3JvdW5kOiR7ZG9jQ29sc1tpXX0iPjwvZGl2PjwvZGl2PjwvZGl2PmA7CiAgfSkuam9pbignJyk7Cn0KCi8vIOKUgOKUgCBEQVkgUExBTiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gcmVuZGVyRGF5UGxhbigpewogIGNvbnN0IGZ1PWZpbHRlcmVkRlUoKTsKICBjb25zdCBmaWx0ZXJLQU09ZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2RwLWthbScpPy52YWx1ZXx8J2FsbCc7CiAgY29uc3QgZmlsdGVyU3JjPWRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdkcC1zcmMnKT8udmFsdWV8fCdhbGwnOwogIGNvbnN0IGZpbHRlckFnPWRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdkcC1hZycpPy52YWx1ZXx8J2FsbCc7CgogIGxldCBpdGVtcz1mdS5maWx0ZXIocj0+CiAgICAoZmlsdGVyS0FNPT09J2FsbCd8fHIua2FtPT09ZmlsdGVyS0FNKSYmCiAgICAoZmlsdGVyU3JjPT09J2FsbCd8fHIuc3JjPT09ZmlsdGVyU3JjKSYmCiAgICAoZmlsdGVyQWc9PT0nYWxsJ3x8ci5hZz09PWZpbHRlckFnfHxyLmFnLmluY2x1ZGVzKGZpbHRlckFnKSkKICApOwoKICAvLyBTdW1tYXJ5IHBpbGxzCiAgY29uc3Qga2Ftcz1bLi4ubmV3IFNldChpdGVtcy5tYXAocj0+ci5rYW0pLmZpbHRlcihCb29sZWFuKSldOwogIGNvbnN0IHRvZGF5PW5ldyBEYXRlKCkudG9JU09TdHJpbmcoKS5zbGljZSgwLDEwKTsKICBjb25zdCBvdmVyZHVlPWl0ZW1zLmZpbHRlcihyPT5yLmZkPHRvZGF5KS5sZW5ndGg7CiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2RwLXN1bW1hcnknKS5pbm5lckhUTUw9WwogICAgYDxzcGFuIGNsYXNzPSJrcGlsbCBwYSI+JHtpdGVtcy5sZW5ndGh9IGZvbGxvdy11cHM8L3NwYW4+YCwKICAgIGA8c3BhbiBjbGFzcz0ia3BpbGwgcHYiPiR7a2Ftcy5sZW5ndGh9IEtBTXM8L3NwYW4+YCwKICAgIG92ZXJkdWU/YDxzcGFuIGNsYXNzPSJrcGlsbCBwciI+JHtvdmVyZHVlfSBvdmVyZHVlPC9zcGFuPmA6JycsCiAgICBgPHNwYW4gY2xhc3M9ImtwaWxsIHBnIj4ke2l0ZW1zLmZpbHRlcihyPT5yLnNyYz09PSdDUk0nKS5sZW5ndGh9IENSTTwvc3Bhbj5gLAogICAgYDxzcGFuIGNsYXNzPSJrcGlsbCBwYiI+JHtpdGVtcy5maWx0ZXIocj0+ci5zcmM9PT0nV0EnKS5sZW5ndGh9IFdoYXRzQXBwPC9zcGFuPmAsCiAgXS5maWx0ZXIoQm9vbGVhbikuam9pbignJyk7CgogIGlmKCFpdGVtcy5sZW5ndGgpewogICAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2RwLWNhcmRzJykuaW5uZXJIVE1MPWA8ZGl2IGNsYXNzPSJkcC1lbXB0eSIgc3R5bGU9ImdyaWQtY29sdW1uOjEvLTEiPk5vIGZvbGxvdy11cHMgZm91bmQgaW4gc2VsZWN0ZWQgZGF0ZSByYW5nZS48YnI+PHNwYW4gc3R5bGU9ImZvbnQtc2l6ZTouNzVyZW07bWFyZ2luLXRvcDo2cHg7ZGlzcGxheTpibG9jaztjb2xvcjp2YXIoLS1kaW0pIj5UcnkgYWRqdXN0aW5nIHRoZSBkYXRlIHJhbmdlIG9yIEtBTSBmaWx0ZXIgYWJvdmUuPC9zcGFuPjwvZGl2PmA7CiAgICByZXR1cm47CiAgfQoKICAvLyBTb3J0OiBvdmVyZHVlIGZpcnN0LCB0aGVuIGJ5IGRhdGUKICBpdGVtcy5zb3J0KChhLGIpPT4oYS5mZHx8JycpPihiLmZkfHwnJyk/MTotMSk7CgogIGNvbnN0IGFnQ29sb3JzPXsKICAgICdDb250cmFjdCBTaWduaW5nJzonYi1jcm0nLCdEb2N1bWVudCBDb2xsZWN0aW9uJzonYi13YScsCiAgICAnTWVldGluZyB3aXRoIGRpZmZlcmVudCBQT0MnOidiLWNvbGQnLCdUbyBleHBsYWluIGFib3V0IG91ciBwbGF0Zm9ybSc6J2Itd2FybScsCiAgfTsKICBjb25zdCBraW5kQ29sb3JzPXsnVGllci0xKEFPVj40MDApJzonYi1ob3QnLCdUaWVyLTIgKEFPViA8NDAwKSc6J2Itd2FybScsJ1RpZXIgMiBBT1YgPCA1MDAnOidiLXdhcm0nLCdUaWVyIDEgQU9WID41MDAnOidiLWhvdCd9OwoKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnZHAtY2FyZHMnKS5pbm5lckhUTUw9aXRlbXMubWFwKHI9PnsKICAgIGNvbnN0IG92ZXJkdWU9ci5mZDx0b2RheTsKICAgIGNvbnN0IGlzVG9kYXk9ci5mZD09PXRvZGF5OwogICAgY29uc3QgYmRDb2xvcj1vdmVyZHVlPyd2YXIoLS1ycyknOmlzVG9kYXk/J3ZhcigtLWFtKSc6J3ZhcigtLWIyKSc7CiAgICBjb25zdCBhZ0Nscz1hZ0NvbG9yc1tyLmFnXXx8J2ItY29sZCc7CiAgICBjb25zdCBraW5kQ2xzPWtpbmRDb2xvcnNbci5raW5kXXx8Jyc7CiAgICBjb25zdCBwaG9uZT1yLnBoPyhyLnBoKycnKS5yZXBsYWNlKC9cLjAkLywnJyk6Jyc7CiAgICBjb25zdCBnbWFwTGluaz1yLmdtYXAmJnIuZ21hcC5zdGFydHNXaXRoKCdodHRwJyk/YDxhIGhyZWY9IiR7ci5nbWFwfSIgdGFyZ2V0PSJfYmxhbmsiPvCfk40gT3BlbiBNYXA8L2E+YDooci5sb2M/YDxzcGFuPvCfk40gJHtyLmxvY308L3NwYW4+YDonJyk7CgogICAgcmV0dXJuIGA8ZGl2IGNsYXNzPSJkcC1jYXJkIiBzdHlsZT0iYm9yZGVyLWNvbG9yOiR7YmRDb2xvcn0yMDtib3JkZXItbGVmdDozcHggc29saWQgJHtiZENvbG9yfSI+CiAgICAgIDxkaXYgY2xhc3M9ImRwLXRvcCI+CiAgICAgICAgPGRpdj4KICAgICAgICAgIDxkaXYgY2xhc3M9ImRwLXJlc3QiPiR7ci5yZXN0fHwn4oCUJ308L2Rpdj4KICAgICAgICAgIDxkaXYgY2xhc3M9ImRwLWthbSI+8J+RpCAke3Iua2FtfHwnVW5rbm93biBLQU0nfTwvZGl2PgogICAgICAgIDwvZGl2PgogICAgICAgIDxkaXYgY2xhc3M9ImRwLWJhZGdlcyI+CiAgICAgICAgICAke292ZXJkdWU/JzxzcGFuIGNsYXNzPSJiYWRnZSBiLWhvdCI+T3ZlcmR1ZTwvc3Bhbj4nOmlzVG9kYXk/JzxzcGFuIGNsYXNzPSJiYWRnZSBiLXdhcm0iPlRvZGF5PC9zcGFuPic6Jyd9CiAgICAgICAgICAke3Iuc3JjPT09J0NSTSc/JzxzcGFuIGNsYXNzPSJiYWRnZSBiLWNybSI+Q1JNPC9zcGFuPic6JzxzcGFuIGNsYXNzPSJiYWRnZSBiLXdhIj5XaGF0c0FwcDwvc3Bhbj4nfQogICAgICAgICAgJHtyLmtpbmQ/YDxzcGFuIGNsYXNzPSJiYWRnZSAke2tpbmRDbHN9Ij4ke3Iua2luZC5yZXBsYWNlKCdUaWVyLTEoQU9WPjQwMCknLCdUaWVyIDEnKS5yZXBsYWNlKCdUaWVyLTIgKEFPViA8NDAwKScsJ1RpZXIgMicpLnJlcGxhY2UoJ1RpZXIgMiBBT1YgPCA1MDAnLCdUaWVyIDInKS5yZXBsYWNlKCdUaWVyIDEgQU9WID41MDAnLCdUaWVyIDEnKX08L3NwYW4+YDonJ30KICAgICAgICA8L2Rpdj4KICAgICAgPC9kaXY+CiAgICAgICR7ci5hZz9gPGRpdiBjbGFzcz0iZHAtcm93Ij48c3BhbiBjbGFzcz0iaWNvIj7wn46vPC9zcGFuPjxzcGFuPkFnZW5kYTogPHNwYW4gY2xhc3M9InZhbCI+JHtyLmFnfTwvc3Bhbj48L3NwYW4+PC9kaXY+YDonJ30KICAgICAgJHtyLnB0cz9gPGRpdiBjbGFzcz0iZHAtcm93Ij48c3BhbiBjbGFzcz0iaWNvIj7wn5OdPC9zcGFuPjxzcGFuIGNsYXNzPSJ2YWwiPiR7ci5wdHN9PC9zcGFuPjwvZGl2PmA6Jyd9CiAgICAgICR7ci5wb2M/YDxkaXYgY2xhc3M9ImRwLXJvdyI+PHNwYW4gY2xhc3M9ImljbyI+8J+RpDwvc3Bhbj48c3Bhbj5QT0M6IDxzcGFuIGNsYXNzPSJ2YWwiPiR7ci5wb2N9PC9zcGFuPiR7cGhvbmU/JyDCtyA8YSBocmVmPSJ0ZWw6JytwaG9uZSsnIj4nK3Bob25lKyc8L2E+JzonJ308L3NwYW4+PC9kaXY+YDonJ30KICAgICAgJHtnbWFwTGluaz9gPGRpdiBjbGFzcz0iZHAtcm93Ij48c3BhbiBjbGFzcz0iaWNvIj7wn5e6PC9zcGFuPjxzcGFuPiR7Z21hcExpbmt9PC9zcGFuPjwvZGl2PmA6Jyd9CiAgICAgIDxkaXYgY2xhc3M9ImRwLXJvdyI+PHNwYW4gY2xhc3M9ImljbyI+8J+ThTwvc3Bhbj48c3Bhbj5Gb2xsb3ctdXA6IDxzcGFuIGNsYXNzPSJ2YWwiPiR7ci5mZH08L3NwYW4+IMK3IExhc3QgdmlzaXQ6ICR7ci52ZH08L3NwYW4+PC9kaXY+CiAgICAgICR7ci5jb20/YDxkaXYgY2xhc3M9ImRwLWNvbW1lbnQiPvCfkqwgIiR7ci5jb219IjwvZGl2PmA6Jyd9CiAgICA8L2Rpdj5gOwogIH0pLmpvaW4oJycpOwp9CgovLyDilIDilIAgUE9BIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApmdW5jdGlvbiByZW5kZXJQT0EoKXsKICBjb25zdCBwb2E9ZmlsdGVyZWRQT0EoKTsKICBjb25zdCBmdT1wb2EuZmlsdGVyKHI9PnIudnR5cGU9PT0nRm9sbG93IHVwJykubGVuZ3RoOwogIGNvbnN0IHdpPXBvYS5maWx0ZXIocj0+ci52dHlwZT09PSdXYWxraW4nKS5sZW5ndGg7CiAgY29uc3Qgb2I9cG9hLmZpbHRlcihyPT5yLmRlPT09J09uYm9hcmRlZCcpLmxlbmd0aDsKCiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ3BvYS1rcGlzJykuaW5uZXJIVE1MPVsKICAgIHtsOidUb3RhbCBQT0EgUmVjb3Jkcycsdjpwb2EubGVuZ3RoLHM6J0luIHJhbmdlJyxjOidrYzEnLHA6JycscGM6J3BiJ30sCiAgICB7bDonRm9sbG93LXVwcycsdjpmdSxzOmAke2ZtdFAoZnUscG9hLmxlbmd0aCl9IG9mIHZpc2l0c2AsYzona2MyJyxwOidQbGFubmVkJyxwYzoncGEnfSwKICAgIHtsOidXYWxrLWlucycsdjp3aSxzOmAke2ZtdFAod2kscG9hLmxlbmd0aCl9IG9mIHZpc2l0c2AsYzona2MzJyxwOidDb2xkIHZpc2l0cycscGM6J3BiJ30sCiAgICB7bDonT25ib2FyZGVkIChFT0QpJyx2Om9iLHM6J0RheSBlbmQgPSBPbmJvYXJkZWQnLGM6J2tjNCcscDpmbXRQKG9iLHBvYS5sZW5ndGgpKycgY29udi4nLHBjOidwZyd9LAogIF0ubWFwKGs9PmA8ZGl2IGNsYXNzPSJrY2FyZCAke2suY30iPjxkaXYgY2xhc3M9ImtsYmwiPiR7ay5sfTwvZGl2PjxkaXYgY2xhc3M9Imt2YWwiPiR7ay52fTwvZGl2PjxkaXYgY2xhc3M9ImtzdWIiPiR7ay5zfTwvZGl2PjxzcGFuIGNsYXNzPSJrcGlsbCAke2sucGN9Ij4ke2sucH08L3NwYW4+PC9kaXY+YCkuam9pbignJyk7CgogIC8vIEJ5IGRhdGUKICBjb25zdCBieURhdGU9e307CiAgcG9hLmZvckVhY2gocj0+ewogICAgaWYoIWJ5RGF0ZVtyLnZkXSlieURhdGVbci52ZF09e2Z1OjAsd2k6MCxvYjowLGZ1cDowLG1kOjAscnM6MH07CiAgICBpZihyLnZ0eXBlPT09J0ZvbGxvdyB1cCcpYnlEYXRlW3IudmRdLmZ1Kys7CiAgICBpZihyLnZ0eXBlPT09J1dhbGtpbicpYnlEYXRlW3IudmRdLndpKys7CiAgICBpZihyLnZ0eXBlPT09J09uYm9hcmRpbmcnKWJ5RGF0ZVtyLnZkXS5vYisrOwogICAgaWYoci5kZT09PSdPbmJvYXJkZWQnKWJ5RGF0ZVtyLnZkXS5vYisrOwogICAgaWYoci5kZT09PSdGb2xsb3cgVXAnKWJ5RGF0ZVtyLnZkXS5mdXArKzsKICAgIGlmKHIuZGU9PT0nTWVldGluZyBEb25lJylieURhdGVbci52ZF0ubWQrKzsKICAgIGlmKHIuZGU9PT0nUmVzY2hlZHVsZWQnKWJ5RGF0ZVtyLnZkXS5ycysrOwogIH0pOwogIGNvbnN0IGRhdGVzPU9iamVjdC5rZXlzKGJ5RGF0ZSkuc29ydCgpOwogIGNvbnN0IGxhYmVscz1kYXRlcy5tYXAoZD0+ZC5zbGljZSg1KSk7CgogIGRjKCdwb2EtZCcpOwogIGNoYXJ0c1sncG9hLWQnXT1uZXcgQ2hhcnQoZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ3BvYS1kYWlseScpLHt0eXBlOidiYXInLGRhdGE6e2xhYmVscyxkYXRhc2V0czpbCiAgICB7bGFiZWw6J0ZvbGxvdy11cCcsZGF0YTpkYXRlcy5tYXAoZD0+YnlEYXRlW2RdLmZ1KSxiYWNrZ3JvdW5kQ29sb3I6JyNmNWE2MjNjYycsYm9yZGVyQ29sb3I6JyNmNWE2MjMnLGJvcmRlcldpZHRoOjEuNSxib3JkZXJSYWRpdXM6NH0sCiAgICB7bGFiZWw6J1dhbGstaW4nLGRhdGE6ZGF0ZXMubWFwKGQ9PmJ5RGF0ZVtkXS53aSksYmFja2dyb3VuZENvbG9yOicjNGE5MGQ5Y2MnLGJvcmRlckNvbG9yOicjNGE5MGQ5Jyxib3JkZXJXaWR0aDoxLjUsYm9yZGVyUmFkaXVzOjR9LAogICAge2xhYmVsOidPbmJvYXJkaW5nJyxkYXRhOmRhdGVzLm1hcChkPT5ieURhdGVbZF0ub2IpLGJhY2tncm91bmRDb2xvcjonIzIyYzU1ZWNjJyxib3JkZXJDb2xvcjonIzIyYzU1ZScsYm9yZGVyV2lkdGg6MS41LGJvcmRlclJhZGl1czo0fSwKICBdfSxvcHRpb25zOntyZXNwb25zaXZlOnRydWUsbWFpbnRhaW5Bc3BlY3RSYXRpbzpmYWxzZSxwbHVnaW5zOntsZWdlbmQ6e2Rpc3BsYXk6ZmFsc2V9fSxzY2FsZXM6e3g6e2dyaWQ6e2Rpc3BsYXk6ZmFsc2V9LHRpY2tzOntjb2xvcjonIzdhODdhOCcsZm9udDp7c2l6ZTo5fSxtYXhSb3RhdGlvbjo0NX19LHk6e2dyaWQ6e2NvbG9yOicjMjQyZDQyJ30sdGlja3M6e2NvbG9yOicjN2E4N2E4Jyxmb250OntzaXplOjEwfX19fX19KTsKCiAgLy8gRGF5IGVuZCBkb251dAogIGNvbnN0IGRlPXtPbmJvYXJkZWQ6cG9hLmZpbHRlcihyPT5yLmRlPT09J09uYm9hcmRlZCcpLmxlbmd0aCwnRm9sbG93IFVwJzpwb2EuZmlsdGVyKHI9PnIuZGU9PT0nRm9sbG93IFVwJykubGVuZ3RoLCdNZWV0aW5nIERvbmUnOnBvYS5maWx0ZXIocj0+ci5kZT09PSdNZWV0aW5nIERvbmUnKS5sZW5ndGgsUmVzY2hlZHVsZWQ6cG9hLmZpbHRlcihyPT5yLmRlPT09J1Jlc2NoZWR1bGVkJykubGVuZ3RofTsKICBjb25zdCBkZUM9WycjMjJjNTVlJywnI2Y1YTYyMycsJyM0YTkwZDknLCcjZjQzZjVlJ107CiAgZGMoJ3BvYS1kZScpOwogIGNoYXJ0c1sncG9hLWRlJ109bmV3IENoYXJ0KGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdwb2EtZGF5ZW5kJykse3R5cGU6J2RvdWdobnV0JyxkYXRhOntsYWJlbHM6T2JqZWN0LmtleXMoZGUpLGRhdGFzZXRzOlt7ZGF0YTpPYmplY3QudmFsdWVzKGRlKSxiYWNrZ3JvdW5kQ29sb3I6ZGVDLm1hcChjPT5jKydjYycpLGJvcmRlckNvbG9yOicjMWEyMDMwJyxib3JkZXJXaWR0aDoyLGhvdmVyT2Zmc2V0OjV9XX0sb3B0aW9uczp7cmVzcG9uc2l2ZTp0cnVlLG1haW50YWluQXNwZWN0UmF0aW86ZmFsc2UsY3V0b3V0Oic2MCUnLHBsdWdpbnM6e2xlZ2VuZDp7ZGlzcGxheTpmYWxzZX19fX0pOwogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdwb2EtZGUtbGVnJykuaW5uZXJIVE1MPU9iamVjdC5rZXlzKGRlKS5tYXAoKGwsaSk9PmA8c3BhbiBjbGFzcz0ibGkiPjxzcGFuIGNsYXNzPSJsc3EiIHN0eWxlPSJiYWNrZ3JvdW5kOiR7ZGVDW2ldfSI+PC9zcGFuPiR7bH06ICR7ZGVbbF19PC9zcGFuPmApLmpvaW4oJycpOwoKICAvLyBUYWJsZQogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdwb2EtdGJsJykuaW5uZXJIVE1MPWA8dGhlYWQ+PHRyPjx0aD5EYXRlPC90aD48dGg+S0FNPC90aD48dGg+UmVzdGF1cmFudDwvdGg+PHRoPlZpc2l0IFR5cGU8L3RoPjx0aD5Mb2NhdGlvbjwvdGg+PHRoPkxlYWQgVHlwZTwvdGg+PHRoPkRheS1FbmQgU3RhdHVzPC90aD48dGg+UmVtYXJrczwvdGg+PC90cj48L3RoZWFkPjx0Ym9keT5gKwogICAgcG9hLm1hcChyPT57CiAgICAgIGNvbnN0IGRlQ2xzPXIuZGU9PT0nT25ib2FyZGVkJz8ncGcnOnIuZGU9PT0nRm9sbG93IFVwJz8ncGEnOnIuZGU9PT0nUmVzY2hlZHVsZWQnPydwcic6J3BiJzsKICAgICAgcmV0dXJuIGA8dHI+PHRkPiR7ci52ZH08L3RkPjx0ZCBzdHlsZT0id2hpdGUtc3BhY2U6bm93cmFwIj4ke3Iua2FtfHwn4oCUJ308L3RkPjx0ZCBzdHlsZT0iZm9udC13ZWlnaHQ6NTAwIj4ke3IucmVzdHx8J+KAlCd9PC90ZD48dGQ+PHNwYW4gY2xhc3M9ImtwaWxsIHBiIj4ke3IudnR5cGV8fCfigJQnfTwvc3Bhbj48L3RkPjx0ZD4ke3IubG9jfHwn4oCUJ308L3RkPjx0ZD4ke3IubGVhZHx8J+KAlCd9PC90ZD48dGQ+JHtyLmRlP2A8c3BhbiBjbGFzcz0ia3BpbGwgJHtkZUNsc30iPiR7ci5kZX08L3NwYW4+YDon4oCUJ308L3RkPjx0ZCBzdHlsZT0iZm9udC1zaXplOi42OHJlbTtjb2xvcjp2YXIoLS1tdSkiPiR7ci5yZW18fCfigJQnfTwvdGQ+PC90cj5gOwogICAgfSkuam9pbignJykrYDwvdGJvZHk+YDsKfQoKLy8g4pSA4pSAIEtBTSBGVU5ORUwg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmZ1bmN0aW9uIHJlbmRlckZ1bm5lbCgpewogIGNvbnN0IGNybT1maWx0ZXJlZENSTSgpOwogIGNvbnN0IFNUQUdFUz1bJ0ludHJvIE1lZXRpbmcnLCdGb2xsb3ctdXAnLCdDb250cmFjdCBTaWduaW5nJywnRG9jIENvbGxlY3Rpb24nLCdPbmJvYXJkZWQnXTsKICBjb25zdCBTVEFHRV9LRVlTPXsnSW50cm8gTWVldGluZyc6J3MtaW50cm8nLCdGb2xsb3ctdXAnOidzLWZ1JywnQ29udHJhY3QgU2lnbmluZyc6J3MtY3MnLCdEb2MgQ29sbGVjdGlvbic6J3MtZGMnLCdPbmJvYXJkZWQnOidzLW9iJ307CgogIGZ1bmN0aW9uIGdldFN0YWdlKHIpewogICAgY29uc3QgbXQ9KHIubXR8fCcnKS50b0xvd2VyQ2FzZSgpLGFnPShyLmFnfHwnJykudG9Mb3dlckNhc2UoKSxvYj0oci5vYnx8JycpLnRvTG93ZXJDYXNlKCk7CiAgICBpZihvYj09PSd5ZXMnKSByZXR1cm4gJ09uYm9hcmRlZCc7CiAgICBpZihhZy5pbmNsdWRlcygnZG9jJykpIHJldHVybiAnRG9jIENvbGxlY3Rpb24nOwogICAgaWYoYWcuaW5jbHVkZXMoJ2NvbnRyYWN0Jyl8fGFnLmluY2x1ZGVzKCdzaWduJyl8fGFnLmluY2x1ZGVzKCdhZ3JlZW1lbnQnKSkgcmV0dXJuICdDb250cmFjdCBTaWduaW5nJzsKICAgIGlmKG10LmluY2x1ZGVzKCdmb2xsb3cnKSkgcmV0dXJuICdGb2xsb3ctdXAnOwogICAgcmV0dXJuICdJbnRybyBNZWV0aW5nJzsKICB9CgogIGNvbnN0IGthbUZ1bm5lbD17fTsKICBjcm0uZm9yRWFjaChyPT57CiAgICBjb25zdCBrPXIua2FtfHwnVW5rbm93bic7CiAgICBpZigha2FtRnVubmVsW2tdKWthbUZ1bm5lbFtrXT17fTsKICAgIGNvbnN0IHM9Z2V0U3RhZ2Uocik7CiAgICBrYW1GdW5uZWxba11bc109KGthbUZ1bm5lbFtrXVtzXXx8MCkrMTsKICB9KTsKCiAgLy8gU29ydCBieSB0b3RhbAogIGNvbnN0IGthbXM9T2JqZWN0LmtleXMoa2FtRnVubmVsKS5zb3J0KChhLGIpPT5PYmplY3QudmFsdWVzKGthbUZ1bm5lbFtiXSkucmVkdWNlKChzLHYpPT5zK3YsMCktT2JqZWN0LnZhbHVlcyhrYW1GdW5uZWxbYV0pLnJlZHVjZSgocyx2KT0+cyt2LDApKTsKCiAgZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2tmLXRhYmxlJykuaW5uZXJIVE1MPQogICAgYDx0aGVhZD48dHI+PHRoPktBTTwvdGg+JHtTVEFHRVMubWFwKHM9PmA8dGggc3R5bGU9InRleHQtYWxpZ246Y2VudGVyIj4ke3N9PC90aD5gKS5qb2luKCcnKX08dGggc3R5bGU9InRleHQtYWxpZ246Y2VudGVyIj5Ub3RhbDwvdGg+PHRoPkNvbnYuIFJhdGU8L3RoPjwvdHI+PC90aGVhZD48dGJvZHk+YCsKICAgIGthbXMubWFwKChrLGkpPT57CiAgICAgIGNvbnN0IGY9a2FtRnVubmVsW2tdOwogICAgICBjb25zdCB0b3RhbD1PYmplY3QudmFsdWVzKGYpLnJlZHVjZSgocyx2KT0+cyt2LDApOwogICAgICBjb25zdCBjb252PWZbJ0ludHJvIE1lZXRpbmcnXT8oZlsnT25ib2FyZGVkJ118fDApL2ZbJ0ludHJvIE1lZXRpbmcnXSoxMDA6MDsKICAgICAgY29uc3QgcmNscz1pPT09MD8ncmsxJzppPT09MT8ncmsyJzppPT09Mj8ncmszJzoncmtuJzsKICAgICAgcmV0dXJuIGA8dHI+PHRkPjxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjZweCI+PHNwYW4gY2xhc3M9InJhbmstYmFkZ2UgJHtyY2xzfSI+JHtpKzF9PC9zcGFuPjxzcGFuIHN0eWxlPSJmb250LXdlaWdodDo2MDAiPiR7a308L3NwYW4+PC9kaXY+PC90ZD5gKwogICAgICAgIFNUQUdFUy5tYXAocz0+YDx0ZCBjbGFzcz0ic3RhZ2UtY2VsbCI+PHNwYW4gY2xhc3M9InN0YWdlLXBpbGwgJHtTVEFHRV9LRVlTW3NdfSI+JHtmW3NdfHwwfTwvc3Bhbj48L3RkPmApLmpvaW4oJycpKwogICAgICAgIGA8dGQgY2xhc3M9InN0YWdlLWNlbGwiIHN0eWxlPSJmb250LXdlaWdodDo3MDA7Y29sb3I6dmFyKC0tdHgpIj4ke3RvdGFsfTwvdGQ+PHRkPjxzcGFuIGNsYXNzPSJrcGlsbCAke2NvbnY+NDA/J3BnJzpjb252PjIwPydwYSc6J3ByJ30iPiR7Y29udi50b0ZpeGVkKDEpfSU8L3NwYW4+PC90ZD48L3RyPmA7CiAgICB9KS5qb2luKCcnKStgPC90Ym9keT5gOwoKICAvLyBTdGFnZSB0b3RhbHMgYmFyCiAgY29uc3QgdG90YWxzPXt9O1NUQUdFUy5mb3JFYWNoKHM9Pnt0b3RhbHNbc109a2Ftcy5yZWR1Y2UoKHN1bSxrKT0+KGthbUZ1bm5lbFtrXSYma2FtRnVubmVsW2tdW3NdP3N1bStrYW1GdW5uZWxba11bc106c3VtKSwwKTt9KTsKICBjb25zdCBtYXhUPU1hdGgubWF4KC4uLk9iamVjdC52YWx1ZXModG90YWxzKSwxKTsKICBjb25zdCBzQ29sb3JzPVsnIzRhOTBkOScsJyNmNWE2MjMnLCcjOGI1Y2Y2JywnIzA2YjZkNCcsJyMyMmM1NWUnXTsKICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnZm4tYmFycycpLmlubmVySFRNTD1TVEFHRVMubWFwKChzLGkpPT4KICAgIGA8ZGl2IGNsYXNzPSJmbi1yb3ciPjxkaXYgY2xhc3M9ImZuLWxhYmVsIj4ke3N9PC9kaXY+PGRpdiBjbGFzcz0iZm4tdHJhY2siPjxkaXYgY2xhc3M9ImZuLWZpbGwiIHN0eWxlPSJ3aWR0aDoke01hdGgucm91bmQodG90YWxzW3NdL21heFQqMTAwKX0lO2JhY2tncm91bmQ6JHtzQ29sb3JzW2ldfSI+JHt0b3RhbHNbc119PC9kaXY+PC9kaXY+PGRpdiBjbGFzcz0iZm4tY291bnQiPiR7dG90YWxzW3NdfTwvZGl2PjwvZGl2PmAKICApLmpvaW4oJycpOwoKICAvLyBDb252ZXJzaW9uIHJhdGUgYmFyIGNoYXJ0CiAgZGMoJ2ZuLWN2Jyk7CiAgY2hhcnRzWydmbi1jdiddPW5ldyBDaGFydChkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnZm4tY29udicpLHt0eXBlOidiYXInLGRhdGE6e2xhYmVsczprYW1zLGRhdGFzZXRzOlt7bGFiZWw6J0NvbnYuICUnLGRhdGE6a2Ftcy5tYXAoaz0+e2NvbnN0IGY9a2FtRnVubmVsW2tdO3JldHVybiBmWydJbnRybyBNZWV0aW5nJ10/KChmWydPbmJvYXJkZWQnXXx8MCkvZlsnSW50cm8gTWVldGluZyddKjEwMCkudG9GaXhlZCgxKTowO30pLGJhY2tncm91bmRDb2xvcjpQQUwubWFwKGM9PmMrJ2JiJyksYm9yZGVyQ29sb3I6UEFMLGJvcmRlcldpZHRoOjEuNSxib3JkZXJSYWRpdXM6NX1dfSxvcHRpb25zOntpbmRleEF4aXM6J3knLHJlc3BvbnNpdmU6dHJ1ZSxtYWludGFpbkFzcGVjdFJhdGlvOmZhbHNlLHBsdWdpbnM6e2xlZ2VuZDp7ZGlzcGxheTpmYWxzZX19LHNjYWxlczp7eDp7bWF4OjEwMCxncmlkOntjb2xvcjonIzI0MmQ0Mid9LHRpY2tzOntjb2xvcjonIzdhODdhOCcsZm9udDp7c2l6ZToxMH0sY2FsbGJhY2s6dj0+disnJSd9fSx5OntncmlkOntkaXNwbGF5OmZhbHNlfSx0aWNrczp7Y29sb3I6JyNkZGUzZjInLGZvbnQ6e3NpemU6MTB9fX19fX0pOwp9CgovLyDilIDilIAgVklTSVRTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApmdW5jdGlvbiByZW5kZXJWaXNpdHMoKXsKICBjb25zdCB3YT1maWx0ZXJlZFdBKCk7CiAgY29uc3Qgd2l0aEZVPXdhLmZpbHRlcihyPT5yLmZkKS5sZW5ndGg7CiAgY29uc3Qgd2l0aEdtYXA9d2EuZmlsdGVyKHI9PnIuZ21hcCYmci5nbWFwLnN0YXJ0c1dpdGgoJ2h0dHAnKSkubGVuZ3RoOwogIGNvbnN0IGthbXM9Wy4uLm5ldyBTZXQod2EubWFwKHI9PnIua2FtKS5maWx0ZXIoQm9vbGVhbikpXS5sZW5ndGg7CgogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCd2bC1rcGlzJykuaW5uZXJIVE1MPVsKICAgIHtsOidUb3RhbCBWaXNpdCBMb2dzJyx2OndhLmxlbmd0aCxzOidXaGF0c0FwcCB0ZWFtIGxvZycsYzona2MxJyxwYzoncGInfSwKICAgIHtsOidXaXRoIEZvbGxvdy11cCBEYXRlJyx2OndpdGhGVSxzOmZtdFAod2l0aEZVLHdhLmxlbmd0aCkrJyBvZiBsb2dzJyxjOidrYzInLHBjOidwYSd9LAogICAge2w6J1dpdGggR01hcCBMaW5rJyx2OndpdGhHbWFwLHM6J0xvY2F0aW9uIGNhcHR1cmVkJyxjOidrYzMnLHBjOidwZyd9LAogICAge2w6J0tBTXMgQWN0aXZlJyx2OmthbXMsczonTG9nZ2VkIHZpc2l0cyBpbiByYW5nZScsYzona2M0JyxwYzoncHYnfSwKICBdLm1hcChrPT5gPGRpdiBjbGFzcz0ia2NhcmQgJHtrLmN9Ij48ZGl2IGNsYXNzPSJrbGJsIj4ke2subH08L2Rpdj48ZGl2IGNsYXNzPSJrdmFsIj4ke2sudn08L2Rpdj48ZGl2IGNsYXNzPSJrc3ViIj4ke2suc308L2Rpdj48c3BhbiBjbGFzcz0ia3BpbGwgJHtrLnBjfSI+PC9zcGFuPjwvZGl2PmApLmpvaW4oJycpOwoKICAvLyBEYWlseQogIGNvbnN0IGJ5RD17fTt3YS5mb3JFYWNoKHI9PntieURbci52ZF09KGJ5RFtyLnZkXXx8MCkrMTt9KTsKICBjb25zdCBkbD1PYmplY3Qua2V5cyhieUQpLnNvcnQoKTsKICBkYygndmwtZCcpOwogIGNoYXJ0c1sndmwtZCddPW5ldyBDaGFydChkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgndmwtZGFpbHknKSx7dHlwZTonYmFyJyxkYXRhOntsYWJlbHM6ZGwubWFwKGQ9PmQuc2xpY2UoNSkpLGRhdGFzZXRzOlt7ZGF0YTpkbC5tYXAoZD0+YnlEW2RdKSxiYWNrZ3JvdW5kQ29sb3I6JyMxNGI4YTY4OCcsYm9yZGVyQ29sb3I6JyMxNGI4YTYnLGJvcmRlcldpZHRoOjEuNSxib3JkZXJSYWRpdXM6NH1dfSxvcHRpb25zOntyZXNwb25zaXZlOnRydWUsbWFpbnRhaW5Bc3BlY3RSYXRpbzpmYWxzZSxwbHVnaW5zOntsZWdlbmQ6e2Rpc3BsYXk6ZmFsc2V9fSxzY2FsZXM6e3g6e2dyaWQ6e2Rpc3BsYXk6ZmFsc2V9LHRpY2tzOntjb2xvcjonIzdhODdhOCcsZm9udDp7c2l6ZTo5fSxtYXhSb3RhdGlvbjo0NX19LHk6e2dyaWQ6e2NvbG9yOicjMjQyZDQyJ30sdGlja3M6e2NvbG9yOicjN2E4N2E4Jyxmb250OntzaXplOjEwfX19fX19KTsKCiAgLy8gS0FNIGJhcgogIGNvbnN0IGttPXt9O3dhLmZvckVhY2gocj0+e2lmKHIua2FtKWttW3Iua2FtXT0oa21bci5rYW1dfHwwKSsxO30pOwogIGNvbnN0IGtzPU9iamVjdC5lbnRyaWVzKGttKS5zb3J0KChhLGIpPT5iWzFdLWFbMV0pOwogIGRjKCd2bC1rJyk7CiAgY2hhcnRzWyd2bC1rJ109bmV3IENoYXJ0KGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCd2bC1rYW0nKSx7dHlwZTonYmFyJyxkYXRhOntsYWJlbHM6a3MubWFwKGs9PmtbMF0pLGRhdGFzZXRzOlt7ZGF0YTprcy5tYXAoaz0+a1sxXSksYmFja2dyb3VuZENvbG9yOlBBTC5tYXAoYz0+YysnYmInKSxib3JkZXJDb2xvcjpQQUwsYm9yZGVyV2lkdGg6MS41LGJvcmRlclJhZGl1czo0fV19LG9wdGlvbnM6e3Jlc3BvbnNpdmU6dHJ1ZSxtYWludGFpbkFzcGVjdFJhdGlvOmZhbHNlLHBsdWdpbnM6e2xlZ2VuZDp7ZGlzcGxheTpmYWxzZX19LHNjYWxlczp7eDp7Z3JpZDp7ZGlzcGxheTpmYWxzZX0sdGlja3M6e2NvbG9yOicjN2E4N2E4Jyxmb250OntzaXplOjl9LG1heFJvdGF0aW9uOjQ1fX0seTp7Z3JpZDp7Y29sb3I6JyMyNDJkNDInfSx0aWNrczp7Y29sb3I6JyM3YTg3YTgnLGZvbnQ6e3NpemU6MTB9fX19fX0pOwoKICAvLyBUYWJsZQogIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCd2bC10YmwnKS5pbm5lckhUTUw9YDx0aGVhZD48dHI+PHRoPkRhdGU8L3RoPjx0aD5SZXN0YXVyYW50PC90aD48dGg+S0FNPC90aD48dGg+TG9jYXRpb248L3RoPjx0aD5QT0M8L3RoPjx0aD5Db250YWN0PC90aD48dGg+Q29tbWVudDwvdGg+PHRoPkZvbGxvdy11cDwvdGg+PHRoPk1hcDwvdGg+PC90cj48L3RoZWFkPjx0Ym9keT5gKwogICAgd2Euc2xpY2UoMCwyMDApLm1hcChyPT57CiAgICAgIGNvbnN0IHBoPShyLnBofHwnJykudG9TdHJpbmcoKS5yZXBsYWNlKC9cLjAkLywnJyk7CiAgICAgIGNvbnN0IG1hcENlbGw9ci5nbWFwJiZyLmdtYXAuc3RhcnRzV2l0aCgnaHR0cCcpP2A8YSBocmVmPSIke3IuZ21hcH0iIHRhcmdldD0iX2JsYW5rIiBzdHlsZT0iY29sb3I6dmFyKC0tYmwpIj7wn5ONPC9hPmA6J+KAlCc7CiAgICAgIHJldHVybiBgPHRyPjx0ZCBzdHlsZT0id2hpdGUtc3BhY2U6bm93cmFwIj4ke3IudmR9PC90ZD48dGQgc3R5bGU9ImZvbnQtd2VpZ2h0OjUwMCI+JHtyLnJlc3R8fCfigJQnfTwvdGQ+PHRkIHN0eWxlPSJ3aGl0ZS1zcGFjZTpub3dyYXA7Zm9udC1zaXplOi43cmVtIj4ke3Iua2FtfHwn4oCUJ308L3RkPjx0ZD4ke3IubG9jfHwn4oCUJ308L3RkPjx0ZD4ke3IucG9jfHwn4oCUJ308L3RkPjx0ZD48YSBocmVmPSJ0ZWw6JHtwaH0iIHN0eWxlPSJjb2xvcjp2YXIoLS1ibCk7Zm9udC1zaXplOi42OHJlbSI+JHtwaHx8J+KAlCd9PC9hPjwvdGQ+PHRkIHN0eWxlPSJmb250LXNpemU6LjdyZW07Y29sb3I6dmFyKC0tbXUpO21heC13aWR0aDoxODBweCI+JHtyLmNvbXx8J+KAlCd9PC90ZD48dGQgc3R5bGU9IndoaXRlLXNwYWNlOm5vd3JhcDtmb250LXNpemU6LjdyZW0iPiR7ci5mZHx8J+KAlCd9PC90ZD48dGQ+JHttYXBDZWxsfTwvdGQ+PC90cj5gOwogICAgfSkuam9pbignJykrYDwvdGJvZHk+YDsKfQoKLy8g4pSA4pSAIEJPT1Qg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmluaXRGaWx0ZXJzKCk7CnNldFByZXNldCgnbXRkJyk7Cjwvc2NyaXB0Pgo8L2JvZHk+CjwvaHRtbD4=").decode('utf-8')

    # Inject data timestamp into topbar
    ts = datetime.now().strftime("%d %b %Y %H:%M")
    tmpl_after_stamped = tmpl_after.replace(
        "CAPS Delivery · Hyderabad",
        f"CAPS Delivery · Updated {ts}"
    )

    html = tmpl_before + data_json + tmpl_after_stamped

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_HTML)
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)

    size_kb = os.path.getsize(out_path) // 1024
    log(f"Dashboard written: {OUTPUT_HTML} ({size_kb} KB)")

    # Also copy to netlify_deploy/index.html if folder exists
    netlify_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "netlify_deploy")
    netlify_index = os.path.join(netlify_dir, "index.html")
    if os.path.exists(netlify_dir):
        import shutil
        shutil.copy2(out_path, netlify_index)
        log(f"Copied to netlify_deploy/index.html → ready to drag & drop!")
    else:
        print("   (netlify_deploy folder not found — skipping auto-copy)")

    # ── Clean up Google Drive temp files ────────────────────────────────────
    if gdrive_tmp and os.path.exists(gdrive_tmp):
        try:
            shutil.rmtree(gdrive_tmp)
        except Exception:
            pass

    # ── AUTO DEPLOY ──────────────────────────────────────────────────────────
    site_url = None

    if GITHUB_REPO_PATH:
        # GitHub → Vercel (primary)
        site_url = push_to_github(out_path, GITHUB_REPO_PATH, GITHUB_COMMIT_MSG)
    elif NETLIFY_SITE_ID and NETLIFY_TOKEN:
        # Netlify (fallback)
        site_url = deploy_netlify(out_path)
    else:
        print("\n" + "─" * 55)
        print("  📋 MANUAL DEPLOY OPTIONS:")
        print()
        print("  Option A — GitHub + Vercel (auto):")
        print("    Set GITHUB_REPO_PATH in this script")
        print()
        print("  Option B — Netlify drag & drop (30 sec):")
        print(f"    1. Open the folder containing {OUTPUT_HTML}")
        print("    2. Go to: https://app.netlify.com/drop")
        print(f"    3. Drag the 'netlify_deploy' folder onto the page")
        print("─" * 55)

    # Open in browser
    print("\n🌐 Opening dashboard in your browser...")
    abs_path = os.path.abspath(out_path)
    if sys.platform == 'darwin':
        # Mac: use 'open' command which respects default browser
        import subprocess as sp
        sp.Popen(['open', abs_path])
    elif sys.platform == 'win32':
        # Windows: use os.startfile
        os.startfile(abs_path)
    else:
        # Linux
        import subprocess as sp
        sp.Popen(['xdg-open', abs_path])
    if site_url:
        print(f"\n🔗 Shareable link: {site_url}")

    print("\n" + "=" * 55)
    print("  Done! Dashboard updated successfully.")
    if LIVE_URL:
        print(f"  🔗 Live at: {LIVE_URL}")
    print("=" * 55)

    # Mac notification
    msg = f"Dashboard updated! Live at {LIVE_URL}" if LIVE_URL else "Dashboard built & pushed to GitHub"
    notify_mac("✅ CAPS Dashboard Updated", msg)

    # Only wait for input if running interactively (not via launchd scheduler)
    if sys.stdin.isatty():
        input("\nPress Enter to close...")

if __name__ == '__main__':
    main()
